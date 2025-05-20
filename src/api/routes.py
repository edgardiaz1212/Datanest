"""
This module takes care of starting the API Server, Loading the DB and Adding the endpoints
"""
from flask import Flask, request, jsonify, url_for, Blueprint, send_file, current_app, g # Ensure g is imported if used elsewhere, not directly here
from api.models import db, UserForm, Equipment, Description, Rack, TrackerUsuario, RegistroDiagnosticoAire, AireAcondicionado,Lectura, Mantenimiento, UmbralConfiguracion, OtroEquipo, Proveedor, ContactoProveedor, ActividadProveedor,  EstatusActividad, DocumentoExterno, OperativaStateEnum, DiagnosticoComponente, TipoAireRelevanteEnum, ParteACEnum
from api.utils import generate_sitemap, APIException
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from sqlalchemy import or_ , func , distinct, desc
from datetime import datetime, timezone, time as time_obj
import traceback
import sys
import os 
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from flask_jwt_extended import create_access_token, jwt_required, get_jwt_identity 
import io
import base64 # Importar base64
from sqlalchemy.orm import aliased # Añadir aliased
from sqlalchemy import Enum as SQLAlchemyEnum
import uuid
import pandas as pd
import json
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl import Workbook

api = Blueprint('api', __name__)

# Allow CORS requests to this API
CORS(api)
# Para formulario 
#manejo de usuarios que usan el formulario
@api.route('/addUser', methods=['POST'])
def addUser():
    if request.method == "POST":
        data_form = request.get_json()
        data = {
            "email": data_form.get('email'),
            "coordination": data_form.get('coordination'),
            "username": data_form.get('username'),
            "clientName": data_form.get('clientName'),
        }
        new_user = UserForm(
            username=data.get('username'),
            coordination=data.get('coordination'),
            email=data.get('email'),
            clientName=data.get('clientName'),
        )
        db.session.add(new_user)
        try:
            db.session.commit()
            user_id = new_user.id
            user_info = {
                "user_id": user_id,
                "username": new_user.username,
                "coordination": new_user.coordination,
                "email": new_user.email,
                "clientName": new_user.clientName,
            }
            return jsonify(user_info), 201
        except Exception as error:
            db.session.rollback()
            return jsonify({"msg": "Error occurred while trying to upload User", "error": str(error)}), 500

@api.route('/users', methods=['GET'])
def get_users():
    try:
        users = UserForm.query.all()
        users_list = [user.serialize() for user in users]
        return jsonify(users_list), 200
    except Exception as error:
        return jsonify({
            "msg": "Error occurred while trying to fetch users",
            "error": str(error)
        }), 500

@api.route('/user/email/<string:email>', methods=['GET'])
def check_email(email):
    user = UserForm.query.filter_by(email=email).first()
    if user:
        return jsonify({"message": "El correo ya está registrado."}), 400
    return jsonify({"message": "Correo disponible."}), 200

@api.route('/user/<int:user_id>', methods=['GET'])
def get_current_user(user_id):
    if request.method == "GET":
        user = UserForm.query.filter_by(id=user_id).first()
        if user:
            user_data = user.serialize()
            return jsonify(user_data), 200
        else:
            return jsonify({"message": "User not found"}), 404

@api.route('/delete_user_data/<int:user_id>' , methods=['DELETE'])
def delete_user_info(user_id):
    users = UserForm.query.get(user_id)
    if not users:
        return jsonify({"msg": "Usuario no encontrado"}), 404
    descriptions = Description.query.filter_by(user_id=user_id).all()
    racks = Rack.query.filter_by(user_id=user_id).all()
    equipments = Equipment.query.filter_by(user_id=user_id).all()
    
    for rack in racks:
        db.session.delete(rack)
    for eq in equipments:
        db.session.delete(eq)
    db.session.delete(users) 
    for desc in descriptions:
        db.session.delete(desc)
   
    try:
        db.session.commit()
        return jsonify({"msg": "Usuario y datos relacionados eliminados correctamente"}), 200
    except Exception as error:
        db.session.rollback()
        return jsonify({"msg": str(error.args)}), 500

#Descripcion base equipo o rack
@api.route('/addDescription', methods=['POST'])
def add_description():
    if request.method == "POST":
        data_form = request.get_json()
        new_description = Description(
            brand=data_form.get('brand'),
            model=data_form.get('model'),
            serial=data_form.get('serial'),
            partNumber=data_form.get('partNumber'),
            five_years_prevition=data_form.get('five_years_prevition'),
            observations=data_form.get('observations'),
            componentType=data_form.get('componentType'),
            requestType=data_form.get('requestType')
        )
        db.session.add(new_description)
        try:
            db.session.commit()
            description_id = new_description.id
            description_info = new_description.serialize()
            return jsonify(description_info), 201
        except Exception as error:
            db.session.rollback()
            return jsonify({"msg": "Error occurred while trying to upload description", "error": str(error)}), 500

@api.route('/addRack', methods=['POST'])
def add_rack():
    if request.method == "POST":
        data_form = request.get_json()

        # Check if user_id is provided in the request
        user_id = data_form.get('user_id')
        if not user_id:
            return jsonify({"msg": "user_id is required"}), 400

        # Validate if the user exists in the database
        user = UserForm.query.get(user_id)
        if not user:
            return jsonify({"msg": "User not found"}), 404

        new_rack = Rack(
            has_cabinet=data_form.get('has_cabinet'),
            leased=data_form.get('leased'),
            total_cabinets=data_form.get('total_cabinets'),
            open_closed=data_form.get('open_closed'),
            security=data_form.get('security'),
            type_security=data_form.get('type_security'),
            has_extractors=data_form.get('has_extractors'),
            extractors_ubication=data_form.get('extractors_ubication'),
            modular=data_form.get('modular'),
            lateral_doors=data_form.get('lateral_doors'),
            lateral_ubication=data_form.get('lateral_ubication'),
            rack_unit=data_form.get('rack_unit'),
            rack_position=data_form.get('rack_position'),
            rack_ubication=data_form.get('rack_ubication'),
            has_accessory=data_form.get('has_accessory'),
            accessory_description=data_form.get('accessory_description'),
            rack_width=data_form.get('rack_width'),
            rack_length=data_form.get('rack_length'),
            rack_height=data_form.get('rack_height'),
            internal_pdu=data_form.get('internal_pdu'),
            input_connector=data_form.get('input_connector'),
            fases=data_form.get('fases'),
            output_connector=data_form.get('output_connector'),
            neutro=data_form.get('neutro'),
            description_id=data_form.get('description_id'),
            user_id=user_id
        )
        db.session.add(new_rack)
        try:
            db.session.commit()
            rack_id = new_rack.id
            rack_info = new_rack.serialize()
            return jsonify(rack_info), 201
        except Exception as error:
            db.session.rollback()
            return jsonify({"msg": "Error occurred while trying to upload Rack", "error": str(error)}), 500

@api.route('/addEquipment', methods=['POST'])
def add_equipment():
    if request.method == "POST":
        data_form = request.get_json()
        new_equipment = Equipment(
            equipment_width=data_form.get('equipment_width'),
            equipment_height=data_form.get('equipment_height'),
            equipment_length=data_form.get('equipment_length'),
            packaging_width=data_form.get('packaging_width'),
            packaging_length=data_form.get('packaging_length'),
            packaging_height=data_form.get('packaging_height'),
            weight=data_form.get('weight'),
            anchor_type=data_form.get('anchor_type'),
            service_area=data_form.get('service_area'),
            service_frontal=data_form.get('service_frontal'),
            service_back=data_form.get('service_back'),
            service_lateral=data_form.get('service_lateral'),
            access_width=data_form.get('access_width'),
            access_inclination=data_form.get('access_inclination'),
            access_length=data_form.get('access_length'),
            rack_number=data_form.get('rack_number'),
            equip_rack_ubication=data_form.get('equip_rack_ubication'),
            rack_unit_position=data_form.get('rack_unit_position'),
            total_rack_units=data_form.get('total_rack_units'),
            ac_dc=data_form.get('ac_dc'),
            input_current=data_form.get('input_current'),
            power=data_form.get('power'),
            power_supply=data_form.get('power_supply'),
            operation_temp=data_form.get('operation_temp'),
            thermal_disipation=data_form.get('thermal_disipation'),
            power_config=data_form.get('power_config'),
            description_id=data_form.get('description_id'),
            user_id=data_form.get('user_id'),
            rack_id=data_form.get('rack_id')
        )
        db.session.add(new_equipment)
        try:
            db.session.commit()
            equipment_id = new_equipment.id
            equipment_info = new_equipment.serialize()
            return jsonify(equipment_info), 201
        except Exception as error:
            db.session.rollback()
            # Log the detailed error message
            import traceback
            error_message = traceback.format_exc()
            print(f"Error occurred while trying to upload Equipment: {error_message}")
            return jsonify({"msg": "Error occurred while trying to upload Equipment", "error": error_message}), 500
#datos rack y equipo
@api.route('/description/<int:user_id>', methods=['GET'])
def get_all_descriptions_by_user(user_id):
    user = UserForm.query.filter_by(id=user_id).first()
    if not user:
        return jsonify({"message": "User not found"}), 404
    
    # Obtener descripciones desde racks
    rack_descriptions = [rack.description for rack in user.racks if rack.description]
    
    # Obtener descripciones desde equipos
    equipment_descriptions = [equipment.description for equipment in user.equipments if equipment.description]
    
    # Combinar y serializar descripciones
    all_descriptions = list(set(rack_descriptions + equipment_descriptions))
    descriptions_data = [description.serialize() for description in all_descriptions]
    
    # Responder según el caso
    if not descriptions_data:
        return jsonify({"message": "No descriptions found for this user"}), 200
    
    return jsonify(descriptions_data), 200

@api.route('/rack/<int:description_id>', methods=['GET'])
def get_rack_by_description(description_id):
    rack = Rack.query.filter_by(description_id=description_id).first()
    if not rack:
        return jsonify({"message": "Rack not found"}), 404
    return jsonify(rack.serialize()), 200

@api.route('/rack', methods=['GET'])
def get_rack():
    try:
        racks = Rack.query.all()
        if not racks:
            return jsonify({"message": "Racks not found"}), 404
        
        serialized_racks = [rack.serialize() for rack in racks]
        return jsonify(serialized_racks), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@api.route('/equipment/<int:description_id>', methods=['GET'])
def get_equipment_by_description(description_id):
    equipment = Equipment.query.filter_by(description_id=description_id).first()
    if not equipment:
        return jsonify({"message": "Equipment not found"}), 404
    return jsonify(equipment.serialize()), 200

@api.route('/descriptions/<int:description_id>', methods=['DELETE'])
def delete_description(description_id):
    description = Description.query.get(description_id)
    if description:
        db.session.delete(description)
        db.session.commit()
        return jsonify({'message': 'Description and related entities deleted successfully'}), 200
    else:
        return jsonify({'message': 'Description not found'}), 404
#actualizaciones equipo y rack
@api.route('/editDescription/<int:description_id>', methods=['PUT'])
def update_description(description_id):
    data_form = request.get_json()
    description = Description.query.get(description_id)
    
    if not description:
        return jsonify({"message": "Description not found"}), 404

    description.brand = data_form.get('brand', description.brand)
    description.model = data_form.get('model', description.model)
    description.serial = data_form.get('serial', description.serial)
    description.partNumber = data_form.get('partNumber', description.partNumber)
    description.five_years_prevition = data_form.get('five_years_prevition', description.five_years_prevition)
    description.observations = data_form.get('observations', description.observations)
    description.componentType = data_form.get('componentType', description.componentType)
    description.requestType = data_form.get('requestType', description.requestType)

    try:
        db.session.commit()
        return jsonify(description.serialize()), 200
    except Exception as error:
        db.session.rollback()
        return jsonify({"msg": "Error occurred while trying to update description", "error": str(error)}), 500

@api.route('/editRack/<int:rack_id>', methods=['PUT'])
def update_rack(rack_id):
    data_form = request.get_json()
    rack = Rack.query.get(rack_id)
    
    if not rack:
        return jsonify({"message": "Rack not found"}), 404

    rack.has_cabinet = data_form.get('has_cabinet', rack.has_cabinet)
    rack.leased = data_form.get('leased', rack.leased)
    rack.total_cabinets = data_form.get('total_cabinets', rack.total_cabinets)
    rack.open_closed = data_form.get('open_closed', rack.open_closed)
    rack.security = data_form.get('security', rack.security)
    rack.type_security = data_form.get('type_security', rack.type_security)
    rack.has_extractors = data_form.get('has_extractors', rack.has_extractors)
    rack.extractors_ubication = data_form.get('extractors_ubication', rack.extractors_ubication)
    rack.modular = data_form.get('modular', rack.modular)
    rack.lateral_doors = data_form.get('lateral_doors', rack.lateral_doors)
    rack.lateral_ubication = data_form.get('lateral_ubication', rack.lateral_ubication)
    rack.rack_unit = data_form.get('rack_unit', rack.rack_unit)
    rack.rack_position = data_form.get('rack_position', rack.rack_position)
    rack.rack_ubication = data_form.get('rack_ubication', rack.rack_ubication)
    rack.has_accessory = data_form.get('has_accessory', rack.has_accessory)
    rack.accessory_description = data_form.get('accessory_description', rack.accessory_description)
    rack.rack_width = data_form.get('rack_width', rack.rack_width)
    rack.rack_length = data_form.get('rack_length', rack.rack_length)
    rack.rack_height = data_form.get('rack_height', rack.rack_height)
    rack.internal_pdu = data_form.get('internal_pdu', rack.internal_pdu)
    rack.input_connector = data_form.get('input_connector', rack.input_connector)
    rack.fases = data_form.get('fases', rack.fases)
    rack.output_connector = data_form.get('output_connector', rack.output_connector)
    rack.neutro = data_form.get('neutro', rack.neutro)

    try:
        db.session.commit()
        return jsonify(rack.serialize()), 200
    except Exception as error:
        db.session.rollback()
        return jsonify({"msg": "Error occurred while trying to update rack", "error": str(error)}), 500

@api.route('/editEquipment/<int:equipment_id>', methods=['PUT'])
def edit_equipment(equipment_id):
    if request.method == 'PUT':
        data_form = request.get_json()
        equipment = Equipment.query.get(equipment_id)
        if not equipment:
            return jsonify({'message': 'Equipment not found'}), 404

        equipment.equipment_width = data_form.get('equipment_width', equipment.equipment_width)
        equipment.equipment_height = data_form.get('equipment_height', equipment.equipment_height)
        equipment.equipment_length = data_form.get('equipment_length', equipment.equipment_length)
        equipment.packaging_width= data_form.get('packaging_width', equipment.packaging_width)
        equipment.packaging_length= data_form.get('packaging_length',equipment.packaging_length)
        equipment.packaging_height= data_form.get('packaging_height',equipment.packaging_height)
        equipment.weight= data_form.get('weight',equipment.weight)
        equipment.anchor_type= data_form.get('anchor_type',equipment.anchor_type)
        equipment.service_area= data_form.get('service_area',equipment.service_area)
        equipment.service_frontal= data_form.get('service_frontal',equipment.service_frontal)
        equipment.service_back= data_form.get('service_back',equipment.service_back)
        equipment.service_lateral= data_form.get('service_lateral',equipment.service_lateral)
        equipment.access_width= data_form.get('access_width',equipment.access_width)
        equipment.access_inclination= data_form.get('access_inclination',equipment.access_inclination)
        equipment.access_length= data_form.get('access_length',equipment.access_length)
        equipment.rack_number= data_form.get('rack_number',equipment.rack_number)
        equipment.equip_rack_ubication= data_form.get('equip_rack_ubication',equipment.equip_rack_ubication)
        equipment.rack_unit_position= data_form.get('rack_unit_position',equipment.rack_unit_position)
        equipment.total_rack_units= data_form.get('total_rack_units',equipment.total_rack_units)
        equipment.ac_dc= data_form.get('ac_dc',equipment.ac_dc)
        equipment.input_current= data_form.get('input_current',equipment.input_current)
        equipment.power= data_form.get('power',equipment.power)
        equipment.power_supply= data_form.get('power_supply',equipment.power_supply)
        equipment.operation_temp= data_form.get('operation_temp',equipment.operation_temp)
        equipment.thermal_disipation= data_form.get('thermal_disipation',equipment.thermal_disipation)
        equipment.power_config= data_form.get('power_config',equipment.power_config)

        try:
            db.session.commit()
            return jsonify(equipment.serialize()), 200
        except Exception as error:
            db.session.rollback()
            return jsonify({"msg": "Error occurred while trying to update Equipment", "error": str(error)}), 500

#Manejo de usuario aircontrol
@api.route('/tracker/register', methods=['POST'])
def register_tracker_user():
    """Registra un nuevo usuario de tipo TrackerUsuario."""
    data_form = request.get_json()
    if not data_form:
        return jsonify({"msg": "No input data provided"}), 400

    # Campos requeridos para TrackerUsuario
    email = data_form.get('email')
    username = data_form.get('username')
    password = data_form.get('password')
    nombre = data_form.get('nombre')
    apellido = data_form.get('apellido')
    rol = data_form.get('rol', 'operador') # Rol por defecto

    if not email or not username or not password or not nombre or not apellido:
        return jsonify({"msg": "Nombre, apellido, email, username, and password are required"}), 400

    # Verificar si ya existe un usuario con ese email o username
    existing_user = TrackerUsuario.query.filter(
        or_(TrackerUsuario.email == email, TrackerUsuario.username == username)
    ).first()
    if existing_user:
        return jsonify({"msg": "Email or username already exists for tracker user"}), 409

    # Crear nuevo TrackerUsuario
    new_tracker_user = TrackerUsuario(
        nombre=nombre,
        apellido=apellido,
        email=email,
        username=username,
        rol=rol,
        activo=True, # Activo por defecto
        fecha_registro=datetime.now(timezone.utc) # Usar UTC es buena práctica
    )
    new_tracker_user.set_password(password) # Hashear y guardar contraseña

    db.session.add(new_tracker_user)
    try:
        db.session.commit()
        return jsonify(new_tracker_user.serialize()), 201
    except Exception as error:
        db.session.rollback()
        print(f"Error registering TrackerUsuario: {str(error)}")
        traceback.print_exc()
        return jsonify({"msg": "Error registering tracker user", "error": str(error)}), 500

@api.route('/tracker/login', methods=['POST'])
def login_tracker_user():
    """Autentica un usuario de tipo TrackerUsuario y devuelve un token JWT."""
    data_form = request.get_json()
    if not data_form:
        return jsonify({"msg": "No input data provided"}), 400

    identifier = data_form.get('identifier') # Puede ser email o username
    password = data_form.get('password')

    if not identifier or not password:
        return jsonify({"msg": "Identifier (email or username) and password are required"}), 400

    # Buscar TrackerUsuario por username o email
    user = TrackerUsuario.query.filter(
        or_(TrackerUsuario.username == identifier, TrackerUsuario.email == identifier)
    ).first()

    # Verificar si el usuario existe, está activo y la contraseña es correcta
    if user and user.activo and user.check_password(password):
        try:
            # Actualizar última conexión
            user.ultima_conexion = datetime.now(timezone.utc)
            db.session.commit()

            access_token = create_access_token(identity=str(user.id)) 

            # --- Devolver el token junto con los datos del usuario ---
            return jsonify({
                "msg": "Tracker login successful",
                "token": access_token, # <--- Token añadido aquí
                "user": user.serialize()
            }), 200

        except Exception as error:
             db.session.rollback()
             print(f"Error during tracker login (DB update or token generation): {str(error)}")
             traceback.print_exc()
             return jsonify({"msg": "An error occurred during tracker login", "error": str(error)}), 500
    else:
        # Mensaje de error genérico por seguridad
        return jsonify({"msg": "Invalid credentials or inactive tracker user"}), 401
    
@api.route('/tracker/users', methods=['GET'])
def get_tracker_users():
    """Obtiene todos los usuarios TrackerUsuario (opcionalmente solo activos)."""
    try:
        # Cambia a False si quieres todos, incluyendo inactivos
        solo_activos = request.args.get('activos', 'true').lower() == 'true'

        query = TrackerUsuario.query
        if solo_activos:
            query = query.filter_by(activo=True)

        users = query.all()
        users_list = [user.serialize() for user in users]
        return jsonify(users_list), 200
    except Exception as error:
        print(f"Error fetching tracker users: {str(error)}")
        traceback.print_exc()
        return jsonify({"msg": "Error fetching tracker users", "error": str(error)}), 500

@api.route('/tracker/user/<int:user_id>', methods=['GET'])
@jwt_required() # <--- Añadido aquí
def get_tracker_user_by_id(user_id):
    """Obtiene un TrackerUsuario específico por su ID. Requiere autenticación."""
    # Opcional: Podrías añadir lógica para verificar roles aquí
    # current_user_id = get_jwt_identity()
    # logged_in_user = TrackerUsuario.query.get(current_user_id)
    # if logged_in_user.rol != 'admin' and current_user_id != user_id:
    #     return jsonify({"msg": "Acceso no autorizado"}), 403

    try:
        user = TrackerUsuario.query.get(user_id)
        if user:
            return jsonify(user.serialize()), 200
        else:
            return jsonify({"msg": "Tracker user not found"}), 404
    except Exception as error:
        print(f"Error fetching tracker user {user_id}: {str(error)}")
        traceback.print_exc()
        return jsonify({"msg": "Error fetching tracker user", "error": str(error)}), 500

@api.route('/tracker/user/<int:user_id>', methods=['PUT'])
@jwt_required() # <--- Añadido aquí
def update_tracker_user(user_id):
    """Actualiza la información de un TrackerUsuario. Requiere autenticación."""
    # --- ¡IMPORTANTE: Añadir verificación de permisos! ---
    current_user_id = get_jwt_identity()
    logged_in_user = TrackerUsuario.query.get(current_user_id)

    # Solo un admin o el propio usuario pueden modificar
    if not logged_in_user or (logged_in_user.rol != 'admin' and current_user_id != user_id):
         return jsonify({"msg": "Acceso no autorizado para modificar este usuario"}), 403
    # --- Fin verificación de permisos ---

    user = TrackerUsuario.query.get(user_id)
    if not user:
        return jsonify({"msg": "Tracker user not found"}), 404

    data_form = request.get_json()
    if not data_form:
        return jsonify({"msg": "No input data provided"}), 400

    updated = False

    # --- Lógica de actualización (sin cambios) ---
    # (Solo permitir que un admin cambie el rol o el estado 'activo')
    is_admin = logged_in_user.rol == 'admin'

    if 'nombre' in data_form and data_form['nombre'] != user.nombre:
        user.nombre = data_form['nombre']
        updated = True
    if 'apellido' in data_form and data_form['apellido'] != user.apellido:
        user.apellido = data_form['apellido']
        updated = True
    if is_admin and 'rol' in data_form and data_form['rol'] != user.rol: # Solo admin cambia rol
        user.rol = data_form['rol']
        updated = True
    if is_admin and 'activo' in data_form and isinstance(data_form['activo'], bool) and data_form['activo'] != user.activo: # Solo admin cambia activo
        user.activo = data_form['activo']
        updated = True
    if 'email' in data_form and data_form['email'] != user.email:
        new_email = data_form['email']
        email_exists = TrackerUsuario.query.filter(TrackerUsuario.email == new_email, TrackerUsuario.id != user_id).first()
        if email_exists:
            return jsonify({"msg": f"Email '{new_email}' is already in use by another tracker user"}), 409
        user.email = new_email
        updated = True
    # --- Fin lógica de actualización ---

    if not updated:
         return jsonify({"msg": "No changes detected for tracker user"}), 200 # O 304

    try:
        db.session.commit()
        return jsonify(user.serialize()), 200
    except Exception as error:
        db.session.rollback()
        print(f"Error updating tracker user {user_id}: {str(error)}")
        traceback.print_exc()
        return jsonify({"msg": "Error updating tracker user", "error": str(error)}), 500

@api.route('/tracker/user/password', methods=['PUT'])
@jwt_required()
def change_tracker_user_password():
    """
    Permite al usuario autenticado cambiar su propia contraseña.
    Requiere autenticación JWT.
    Recibe 'current_password' y 'new_password' en el cuerpo JSON.
    """
    current_user_id_str = get_jwt_identity() # El identity es un string
    try:
        current_user_id = int(current_user_id_str)
    except ValueError:
        return jsonify({"msg": "Identidad de usuario inválida en el token"}), 400

    user = db.session.get(TrackerUsuario, current_user_id)
    if not user:
        # Esto no debería pasar si el token es válido, pero por seguridad
        return jsonify({"msg": "Usuario no encontrado"}), 404

    data = request.get_json()
    if not data:
        return jsonify({"msg": "No se recibieron datos JSON"}), 400

    current_password = data.get('current_password')
    new_password = data.get('new_password')

    if not current_password or not new_password:
        return jsonify({"msg": "Se requieren 'current_password' y 'new_password'"}), 400

    # Verificar la contraseña actual
    if not user.check_password(current_password):
        return jsonify({"msg": "La contraseña actual es incorrecta"}), 401 # Unauthorized o Bad Request

    # Validar longitud mínima de la nueva contraseña (opcional pero recomendado)
    if len(new_password) < 6:
         return jsonify({"msg": "La nueva contraseña debe tener al menos 6 caracteres"}), 400

    # Establecer y hashear la nueva contraseña
    try:
        user.set_password(new_password)
        db.session.commit()
        return jsonify({"msg": "Contraseña actualizada correctamente"}), 200
    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"!!! ERROR SQLAlchemy al cambiar contraseña para usuario {current_user_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error de base de datos al actualizar la contraseña."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"!!! ERROR inesperado al cambiar contraseña para usuario {current_user_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor al actualizar la contraseña."}), 500


@api.route('/tracker/user/<int:user_id>', methods=['DELETE'])
@jwt_required() # <--- Añadido aquí
def delete_tracker_user(user_id):
    """Elimina un TrackerUsuario. Requiere autenticación y permisos de admin."""
    # --- ¡IMPORTANTE: Añadir verificación de permisos! ---
    current_user_id = get_jwt_identity()
    logged_in_user = TrackerUsuario.query.get(current_user_id)

    # Solo un admin puede eliminar usuarios
    if not logged_in_user or logged_in_user.rol != 'admin':
         return jsonify({"msg": "Acceso no autorizado para eliminar usuarios"}), 403

    # Evitar que un admin se elimine a sí mismo (opcional pero recomendado)
    if current_user_id == user_id:
        return jsonify({"msg": "No puedes eliminar tu propia cuenta de administrador"}), 403
    # --- Fin verificación de permisos ---

    user = TrackerUsuario.query.get(user_id)
    if not user:
        return jsonify({"msg": "Tracker user not found"}), 404

    db.session.delete(user)
    try:
        db.session.commit()
        return jsonify({"msg": f"Tracker user {user_id} deleted successfully"}), 200
    except Exception as error:
        db.session.rollback()
        print(f"Error deleting tracker user {user_id}: {str(error)}")
        traceback.print_exc()
        return jsonify({"msg": "Error deleting tracker user", "error": str(error)}), 500

# --- Rutas para AireAcondicionado ---

@api.route('/aires', methods=['POST'])
@jwt_required() # <--- Añadido aquí
def agregar_aire_route():
    """
    Endpoint para agregar un nuevo aire acondicionado. Requiere autenticación.
    Recibe los datos en formato JSON.
    """
    # Opcional: Verificar si el usuario tiene permiso para agregar (ej: rol admin o supervisor)
    # current_user_id = get_jwt_identity()
    # logged_in_user = TrackerUsuario.query.get(current_user_id)
    # if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor']:
    #     return jsonify({"msg": "Acceso no autorizado para agregar aires"}), 403

    data = request.get_json()
    if not data:
        return jsonify({"msg": "No se recibieron datos JSON"}), 400

    # (Resto del código de la función sin cambios...)
    required_fields = [
        'nombre', 'ubicacion', 'fecha_instalacion', 'tipo', 'toneladas',
        'evaporadora_operativa', 'evaporadora_marca', 'evaporadora_modelo', 'evaporadora_serial',
        'evaporadora_codigo_inventario', 'evaporadora_ubicacion_instalacion', 'condensadora_operativa', 'condensadora_marca', 'condensadora_modelo', 'condensadora_serial',
        'condensadora_codigo_inventario', 'condensadora_ubicacion_instalacion'
    ]
    if not all(field in data for field in required_fields):
        missing = [field for field in required_fields if field not in data]
        return jsonify({"msg": f"Faltan campos requeridos: {', '.join(missing)}"}), 400

    try:
        fecha_instalacion_dt = None
        if data.get('fecha_instalacion'):
            try:
                fecha_instalacion_dt = datetime.strptime(data['fecha_instalacion'], '%Y-%m-%d').date()
            except ValueError:
                return jsonify({"msg": "Formato de fecha_instalacion inválido. Usar YYYY-MM-DD."}), 400

        toneladas_float = None
        if data.get('toneladas') is not None:
             try:
                 toneladas_float = float(data['toneladas'])
             except (ValueError, TypeError):
                 return jsonify({"msg": "Valor de toneladas inválido. Debe ser numérico."}), 400

        # Validar y convertir estados operativos
        try:
            evaporadora_operativa_enum = OperativaStateEnum(data['evaporadora_operativa'])
            condensadora_operativa_enum = OperativaStateEnum(data['condensadora_operativa'])
        except ValueError:
            valid_states = [s.value for s in OperativaStateEnum]
            return jsonify({"msg": f"Valores inválidos para estados operativos. Usar: {', '.join(valid_states)}"}), 400
        except KeyError:
            return jsonify({"msg": "Faltan campos 'evaporadora_operativa' o 'condensadora_operativa'."}), 400

        nuevo_aire = AireAcondicionado(
            nombre=data['nombre'],
            ubicacion=data['ubicacion'],
            fecha_instalacion=fecha_instalacion_dt,
            tipo=data['tipo'],
            toneladas=toneladas_float,
            evaporadora_operativa=evaporadora_operativa_enum,
            evaporadora_marca=data['evaporadora_marca'],
            evaporadora_modelo=data['evaporadora_modelo'],
            evaporadora_serial=data['evaporadora_serial'],
            evaporadora_codigo_inventario=data['evaporadora_codigo_inventario'],
            evaporadora_ubicacion_instalacion=data['evaporadora_ubicacion_instalacion'],
            # Los campos de diagnóstico ya NO se guardan directamente aquí
            # evaporadora_diagnostico_id=data.get('evaporadora_diagnostico_id'),
            # evaporadora_diagnostico_notas=data.get('evaporadora_diagnostico_notas'),
            # evaporadora_fecha_hora_diagnostico=evap_fecha_hora_diagnostico_dt,
            condensadora_operativa=condensadora_operativa_enum,
            condensadora_marca=data['condensadora_marca'],
            condensadora_modelo=data['condensadora_modelo'],
            condensadora_serial=data['condensadora_serial'],
            condensadora_codigo_inventario=data['condensadora_codigo_inventario'],
            condensadora_ubicacion_instalacion=data['condensadora_ubicacion_instalacion']
        )
        # Los campos de diagnóstico ya NO se guardan directamente aquí
        # condensadora_diagnostico_id=data.get('condensadora_diagnostico_id'),
        # condensadora_diagnostico_notas=data.get('condensadora_diagnostico_notas'),

        db.session.add(nuevo_aire)
        db.session.commit()

        return jsonify(nuevo_aire.serialize()), 201

    except IntegrityError as e:
        db.session.rollback()
        error_info = str(e.orig)
        msg = "Error: Ya existe un registro con ese Serial o Código de Inventario."
        if 'UNIQUE constraint failed' in error_info:
             if 'evaporadora_serial' in error_info or 'condensadora_serial' in error_info:
                 msg = "Error: Ya existe un aire con ese número de serie."
             elif 'evaporadora_codigo_inventario' in error_info or 'condensadora_codigo_inventario' in error_info:
                 msg = "Error: Ya existe un aire con ese código de inventario."
        print(f"Error de integridad al agregar aire: {e}", file=sys.stderr)
        return jsonify({"msg": msg}), 409

    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"!!! ERROR SQLAlchemy en agregar_aire_route: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error de base de datos al agregar el aire."}), 500

    except Exception as e:
        db.session.rollback()
        print(f"!!! ERROR inesperado en agregar_aire_route: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor."}), 500

@api.route('/aires', methods=['GET'])
@jwt_required() 
def obtener_aires_route():
    """
    Endpoint para obtener la lista de todos los aires acondicionados. Requiere autenticación.
    """
    # No se necesita get_jwt_identity() aquí a menos que quieras filtrar por usuario
    try:
        aires = AireAcondicionado.query.order_by(AireAcondicionado.nombre).all()
        aires_serializados = [aire.serialize() for aire in aires]
        return jsonify(aires_serializados), 200
    except Exception as e:
        print(f"!!! ERROR inesperado en obtener_aires_route: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor al obtener aires."}), 500

@api.route('/aires/<int:aire_id>', methods=['GET'])
@jwt_required() 
def obtener_aire_por_id_route(aire_id):
    """
    Endpoint para obtener un aire acondicionado específico por su ID. Requiere autenticación.
    """
    # No se necesita get_jwt_identity() aquí a menos que quieras verificar permisos específicos
    try:
        if aire_id <= 0:
             return jsonify({"msg": "ID de aire inválido."}), 400

        aire = db.session.get(AireAcondicionado, aire_id)

        if not aire:
            return jsonify({"msg": f"Aire acondicionado con ID {aire_id} no encontrado."}), 404

        return jsonify(aire.serialize()), 200

    except Exception as e:
        print(f"!!! ERROR inesperado en obtener_aire_por_id_route para ID {aire_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor al obtener el aire."}), 500

@api.route('/aires/<int:aire_id>', methods=['PUT'])
@jwt_required() # <--- Añadido aquí
def actualizar_aire_route(aire_id):

    current_user_id = get_jwt_identity()
    logged_in_user = TrackerUsuario.query.get(current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor']:
        return jsonify({"msg": "Acceso no autorizado para actualizar aires"}), 403

    aire = AireAcondicionado.query.get(aire_id)
    if not aire:
        return jsonify({"msg": f"Aire acondicionado con ID {aire_id} no encontrado."}), 404

    data = request.get_json()
    if not data:
        return jsonify({"msg": "No se recibieron datos JSON"}), 400
    # Guardar estados anteriores ANTES de modificar el objeto 'aire'
    estado_anterior_evap = aire.evaporadora_operativa
    estado_anterior_cond = aire.condensadora_operativa

    # (Resto del código de la función sin cambios...)
    try:
        aire.nombre = data.get('nombre', aire.nombre)
        aire.ubicacion = data.get('ubicacion', aire.ubicacion)
        aire.tipo = data.get('tipo', aire.tipo)

        if 'fecha_instalacion' in data:
            fecha_str = data['fecha_instalacion']
            if fecha_str:
                try:
                    aire.fecha_instalacion = datetime.strptime(fecha_str, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    return jsonify({"msg": "Formato de fecha_instalacion inválido. Usar YYYY-MM-DD o null."}), 400
            else:
                 aire.fecha_instalacion = None

        if 'toneladas' in data:
            toneladas_val = data['toneladas']
            if toneladas_val is not None and toneladas_val != '':
                 try:
                     aire.toneladas = float(toneladas_val)
                 except (ValueError, TypeError):
                     return jsonify({"msg": "Valor de toneladas inválido. Debe ser numérico."}), 400
            else:
                 aire.toneladas = None

        if 'evaporadora_operativa' in data:
            try:
                nuevo_estado_evap_enum = OperativaStateEnum(data['evaporadora_operativa'])
                aire.evaporadora_operativa = nuevo_estado_evap_enum # Aplicar el cambio
                # Si cambió A operativa DESDE un estado no operativo
                if nuevo_estado_evap_enum == OperativaStateEnum.OPERATIVA and estado_anterior_evap != OperativaStateEnum.OPERATIVA:
                    marcar_diagnosticos_como_solucionados(aire_id, ParteACEnum.EVAPORADORA, db.session)
            except ValueError:
                valid_states = [s.value for s in OperativaStateEnum]
                return jsonify({"msg": f"Valor inválido para evaporadora_operativa. Usar: {', '.join(valid_states)}"}), 400
        
        aire.evaporadora_marca = data.get('evaporadora_marca', aire.evaporadora_marca)
        aire.evaporadora_modelo = data.get('evaporadora_modelo', aire.evaporadora_modelo)
        aire.evaporadora_serial = data.get('evaporadora_serial', aire.evaporadora_serial)
        aire.evaporadora_codigo_inventario = data.get('evaporadora_codigo_inventario', aire.evaporadora_codigo_inventario)
        aire.evaporadora_ubicacion_instalacion = data.get('evaporadora_ubicacion_instalacion', aire.evaporadora_ubicacion_instalacion)

        if 'condensadora_operativa' in data:
            try:
                nuevo_estado_cond_enum = OperativaStateEnum(data['condensadora_operativa'])
                aire.condensadora_operativa = nuevo_estado_cond_enum # Aplicar el cambio
                if nuevo_estado_cond_enum == OperativaStateEnum.OPERATIVA and estado_anterior_cond != OperativaStateEnum.OPERATIVA:
                    marcar_diagnosticos_como_solucionados(aire_id, ParteACEnum.CONDENSADORA, db.session)
            except ValueError:
                valid_states = [s.value for s in OperativaStateEnum]
                return jsonify({"msg": f"Valor inválido para condensadora_operativa. Usar: {', '.join(valid_states)}"}), 400

        aire.condensadora_marca = data.get('condensadora_marca', aire.condensadora_marca)
        aire.condensadora_modelo = data.get('condensadora_modelo', aire.condensadora_modelo)
        aire.condensadora_serial = data.get('condensadora_serial', aire.condensadora_serial)
        aire.condensadora_codigo_inventario = data.get('condensadora_codigo_inventario', aire.condensadora_codigo_inventario)
        aire.condensadora_ubicacion_instalacion = data.get('condensadora_ubicacion_instalacion', aire.condensadora_ubicacion_instalacion)

        db.session.commit()
        return jsonify(aire.serialize()), 200

    except IntegrityError as e:
        db.session.rollback()
        error_info = str(e.orig)
        msg = "Error: Ya existe otro registro con ese Serial o Código de Inventario."
        print(f"Error de integridad al actualizar aire {aire_id}: {e}", file=sys.stderr)
        return jsonify({"msg": msg}), 409

    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"!!! ERROR SQLAlchemy al actualizar aire ID {aire_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error de base de datos al actualizar el aire."}), 500

    except Exception as e:
        db.session.rollback()
        print(f"!!! ERROR inesperado al actualizar aire ID {aire_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor."}), 500

@api.route('/aires/<int:aire_id>', methods=['DELETE'])
@jwt_required() 
def eliminar_aire_route(aire_id):
    """
    Endpoint para eliminar un aire acondicionado específico por su ID. Requiere autenticación.
    """

    # Solo un admin o supervisor debería poder eliminar
    current_user_id = get_jwt_identity()
    logged_in_user = TrackerUsuario.query.get(current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor']:
         return jsonify({"msg": "Acceso no autorizado para eliminar aires"}), 403

    try:
        aire = db.session.get(AireAcondicionado, aire_id)

        if not aire:
            return jsonify({"msg": f"Aire acondicionado con ID {aire_id} no encontrado."}), 404

        db.session.delete(aire)
        db.session.commit()

        return '', 204

    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"!!! ERROR SQLAlchemy al eliminar aire ID {aire_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error de base de datos al eliminar el aire."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"!!! ERROR inesperado al eliminar aire ID {aire_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor al eliminar el aire."}), 500

# --- Rutas para DiagnosticoComponente ---
def marcar_diagnosticos_como_solucionados(aire_id, parte_afectada_enum, session):
    """
    Marca los registros de diagnóstico no solucionados para una parte específica de un aire como solucionados.
    'session' es la sesión de SQLAlchemy (db.session).
    """
    try:
        registros_a_solucionar = session.query(RegistroDiagnosticoAire).filter(
            RegistroDiagnosticoAire.aire_id == aire_id,
            RegistroDiagnosticoAire.parte_ac == parte_afectada_enum,
            RegistroDiagnosticoAire.solucionado == False  # Solo los no solucionados
        ).all()

        if registros_a_solucionar:
            print(f"DEBUG: Marcando {len(registros_a_solucionar)} diagnósticos como solucionados para aire {aire_id}, parte {parte_afectada_enum.value}", file=sys.stderr)
            for registro in registros_a_solucionar:
                registro.solucionado = True
                registro.fecha_solucion = datetime.now(timezone.utc)
            # El commit se hará fuera de esta función, en la ruta que llama a este helper
            # session.commit() # No hacer commit aquí
        else:
            print(f"DEBUG: No hay diagnósticos pendientes para marcar como solucionados para aire {aire_id}, parte {parte_afectada_enum.value}", file=sys.stderr)
        return True # Indicar éxito
    except Exception as e:
        print(f"ERROR: Excepción al marcar diagnósticos como solucionados para aire {aire_id}, parte {parte_afectada_enum.value}: {e}", file=sys.stderr)
        # No hacer rollback aquí, dejar que la función llamante lo maneje.
        # session.rollback() # No hacer rollback aquí
        # raise # Re-lanzar para que la transacción principal falle si es necesario
        return False # Indicar fallo
    
@api.route('/diagnostico_componentes', methods=['POST'])
@jwt_required()
def crear_diagnostico_componente_route():
    """
    Endpoint para crear un nuevo diagnóstico predefinido. Requiere autenticación (Admin/Supervisor).
    Recibe los datos en formato JSON.
    """
    current_user_id = get_jwt_identity()
    logged_in_user = TrackerUsuario.query.get(current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor']:
        return jsonify({"msg": "Acceso no autorizado para crear diagnósticos"}), 403

    data = request.get_json()
    if not data:
        return jsonify({"msg": "No se recibieron datos JSON"}), 400

    required_fields = ['nombre', 'parte_ac', 'tipo_aire_sugerido']
    if not all(field in data for field in required_fields):
        missing = [field for field in required_fields if field not in data]
        return jsonify({"msg": f"Faltan campos requeridos: {', '.join(missing)}"}), 400

    try:
        # Validar y convertir enums
        try:
            parte_ac_enum = ParteACEnum(data['parte_ac'])
            tipo_aire_sugerido_enum = TipoAireRelevanteEnum(data['tipo_aire_sugerido'])
        except ValueError:
            return jsonify({"msg": "Valores inválidos para 'parte_ac' o 'tipo_aire_sugerido'."}), 400

        nuevo_diagnostico = DiagnosticoComponente(
            nombre=data['nombre'],
            parte_ac=parte_ac_enum,
            tipo_aire_sugerido=tipo_aire_sugerido_enum,
            descripcion_ayuda=data.get('descripcion_ayuda'),
            activo=bool(data.get('activo', True)) # Default a True si no se especifica
        )

        db.session.add(nuevo_diagnostico)
        db.session.commit()

        return jsonify(nuevo_diagnostico.serialize()), 201

    except IntegrityError:
        db.session.rollback()
        return jsonify({"msg": f"Ya existe un diagnóstico con el nombre '{data['nombre']}'."}), 409
    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"!!! ERROR SQLAlchemy en crear_diagnostico_componente_route: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error de base de datos al crear el diagnóstico."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"!!! ERROR inesperado en crear_diagnostico_componente_route: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor."}), 500

@api.route('/diagnostico_componentes', methods=['GET'])
@jwt_required()
def obtener_diagnosticos_componentes_route():
    """
    Endpoint para obtener la lista de diagnósticos predefinidos. Requiere autenticación.
    Opcionalmente filtra por ?activo=true/false.
    """
    try:
        query = db.session.query(DiagnosticoComponente)

        activo_filter = request.args.get('activo')
        if activo_filter is not None:
            try:
                activo_bool = activo_filter.lower() == 'true'
                query = query.filter(DiagnosticoComponente.activo == activo_bool)
            except ValueError:
                 return jsonify({"msg": "Valor inválido para el filtro 'activo'. Usar 'true' o 'false'."}), 400

        diagnosticos = query.order_by(DiagnosticoComponente.nombre).all()
        results = [d.serialize() for d in diagnosticos]

        return jsonify(results), 200

    except Exception as e:
        print(f"!!! ERROR inesperado en obtener_diagnosticos_componentes_route: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor al obtener diagnósticos."}), 500

@api.route('/diagnostico_componentes/<int:diagnostico_id>', methods=['GET'])
@jwt_required()
def obtener_diagnostico_componente_por_id_route(diagnostico_id):
    """Obtiene un diagnóstico predefinido específico por su ID. Requiere autenticación."""
    try:
        diagnostico = db.session.get(DiagnosticoComponente, diagnostico_id)
        if not diagnostico:
            return jsonify({"msg": f"Diagnóstico con ID {diagnostico_id} no encontrado."}), 404
        return jsonify(diagnostico.serialize()), 200
    except Exception as e:
        print(f"!!! ERROR inesperado en obtener_diagnostico_componente_por_id_route para ID {diagnostico_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor al obtener el diagnóstico."}), 500

# --- Rutas para RegistroDiagnosticoAire ---

@api.route('/aires/<int:aire_id>/registros_diagnostico', methods=['POST'])
@jwt_required()
def crear_registro_diagnostico_aire_route(aire_id):
    """
    Endpoint para agregar un nuevo registro de diagnóstico histórico para un aire acondicionado.
    Requiere autenticación (Admin/Supervisor/Tecnico).
    Recibe los datos en formato JSON.
    """
    current_user_id = get_jwt_identity()
    logged_in_user = TrackerUsuario.query.get(current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor', 'tecnico']:
        return jsonify({"msg": "Acceso no autorizado para crear registros de diagnóstico"}), 403

    aire = db.session.get(AireAcondicionado, aire_id)
    if not aire:
        return jsonify({"msg": f"Aire acondicionado con ID {aire_id} no encontrado."}), 404

    data = request.get_json()
    if not data:
        return jsonify({"msg": "No se recibieron datos JSON"}), 400

    required_fields = ['parte_ac', 'diagnostico_id', 'fecha_hora']
    if not all(field in data for field in required_fields):
        missing = [field for field in required_fields if field not in data]
        return jsonify({"msg": f"Faltan campos requeridos: {', '.join(missing)}"}), 400

    try:
        # Validar y convertir enums
        try:
            parte_ac_enum = ParteACEnum(data['parte_ac'])
        except ValueError:
            valid_parts = [p.value for p in ParteACEnum]
            return jsonify({"msg": f"Valor inválido para 'parte_ac'. Usar: {', '.join(valid_parts)}"}), 400

        # Validar que el diagnostico_id exista
        diagnostico = db.session.get(DiagnosticoComponente, data['diagnostico_id'])
        if not diagnostico:
            return jsonify({"msg": f"Diagnóstico predefinido con ID {data['diagnostico_id']} no encontrado."}), 404

        # Convertir fecha_hora de string a datetime
        fecha_hora_dt = None
        try:
            fecha_hora_dt = datetime.fromisoformat(data['fecha_hora'])
        except (ValueError, TypeError):
            return jsonify({"msg": "Formato de fecha_hora inválido. Usar formato ISO 8601 (YYYY-MM-DDTHH:MM:SS)."}), 400

        nuevo_registro = RegistroDiagnosticoAire(
            aire_id=aire_id,
            parte_ac=parte_ac_enum,
            diagnostico_id=data['diagnostico_id'],
            fecha_hora=fecha_hora_dt,
            notas=data.get('notas'),
            registrado_por_usuario_id=current_user_id # Registrar quién creó el registro
        )

        db.session.add(nuevo_registro)
        db.session.commit()

        return jsonify(nuevo_registro.serialize()), 201

    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"!!! ERROR SQLAlchemy en crear_registro_diagnostico_aire_route para aire {aire_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error de base de datos al agregar el registro de diagnóstico."}), 500

    except Exception as e:
        db.session.rollback()
        print(f"!!! ERROR inesperado en crear_registro_diagnostico_aire_route para aire {aire_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor."}), 500

@api.route('/aires/<int:aire_id>/registros_diagnostico', methods=['GET'])
@jwt_required()
def obtener_registros_diagnostico_por_aire_route(aire_id):
    """
    Endpoint para obtener todos los registros de diagnóstico histórico de un aire acondicionado específico.
    Requiere autenticación.
    """
    # Verificar que el aire acondicionado existe (opcional, la query ya lo filtra)
    # aire = db.session.get(AireAcondicionado, aire_id)
    # if not aire:
    #     return jsonify({"msg": f"Aire acondicionado con ID {aire_id} no encontrado."}), 404

    try:
        # Obtener registros ordenados por fecha descendente
        registros = db.session.query(RegistroDiagnosticoAire)\
            .filter_by(aire_id=aire_id)\
            .order_by(RegistroDiagnosticoAire.fecha_hora.desc())\
            .all()

        results = [r.serialize() for r in registros]
        return jsonify(results), 200

    except Exception as e:
        print(f"!!! ERROR inesperado en obtener_registros_diagnostico_por_aire_route para aire {aire_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor al obtener registros de diagnóstico."}), 500

@api.route('/registros_diagnostico/todos', methods=['GET'])
@jwt_required()
def obtener_todos_registros_diagnostico_route():
    """
    Endpoint para obtener todos los registros de diagnóstico,
    opcionalmente filtrados por 'solucionado'.
    Requiere autenticación.
    """
    try:
        query = db.session.query(RegistroDiagnosticoAire)

        solucionado_filter_str = request.args.get('solucionado')
        if solucionado_filter_str is not None:
            solucionado_filter_bool = solucionado_filter_str.lower() == 'true'
            query = query.filter(RegistroDiagnosticoAire.solucionado == solucionado_filter_bool)

        # Eager load related data to avoid N+1 queries if serializing related objects
        query = query.options(
            db.joinedload(RegistroDiagnosticoAire.aire), # Para tener aire.nombre, aire.ubicacion
            db.joinedload(RegistroDiagnosticoAire.diagnostico),
            db.joinedload(RegistroDiagnosticoAire.registrado_por)
        )

        registros = query.order_by(RegistroDiagnosticoAire.fecha_hora.desc()).all()
        
        # Modificar la serialización para incluir nombre y ubicación del aire
        results = []
        for r in registros:
            serialized_record = r.serialize()
            if r.aire: # Asegurarse que la relación aire está cargada
                serialized_record['aire_nombre'] = r.aire.nombre
                serialized_record['aire_ubicacion'] = r.aire.ubicacion
            else:
                serialized_record['aire_nombre'] = "N/A"
                serialized_record['aire_ubicacion'] = "N/A"
            results.append(serialized_record)

        return jsonify(results), 200

    except Exception as e:
        print(f"!!! ERROR inesperado en obtener_todos_registros_diagnostico_route: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor al obtener todos los registros de diagnóstico."}), 500


@api.route('/registros_diagnostico/<int:registro_id>', methods=['PUT'])
@jwt_required()
def actualizar_registro_diagnostico_aire_route(registro_id):
    """
    Endpoint para actualizar un registro de diagnóstico histórico existente.
    Requiere autenticación (Admin/Supervisor/Tecnico).
    Recibe los datos en formato JSON.
    """
    current_user_id = get_jwt_identity()
    logged_in_user = TrackerUsuario.query.get(current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor', 'tecnico']:
        return jsonify({"msg": "Acceso no autorizado para actualizar registros de diagnóstico"}), 403

    registro = db.session.get(RegistroDiagnosticoAire, registro_id)
    if not registro:
        return jsonify({"msg": f"Registro de diagnóstico con ID {registro_id} no encontrado."}), 404

    data = request.get_json()
    if not data:
        return jsonify({"msg": "No se recibieron datos JSON"}), 400

    updated = False
    try:
        # Solo permitir actualizar campos específicos
        if 'parte_ac' in data and data['parte_ac'] != registro.parte_ac.value:
            try:
                registro.parte_ac = ParteACEnum(data['parte_ac'])
                updated = True
            except ValueError:
                valid_parts = [p.value for p in ParteACEnum]
                return jsonify({"msg": f"Valor inválido para 'parte_ac'. Usar: {', '.join(valid_parts)}"}), 400

        if 'diagnostico_id' in data and data['diagnostico_id'] != registro.diagnostico_id:
            # Validar que el nuevo diagnostico_id exista
            diagnostico = db.session.get(DiagnosticoComponente, data['diagnostico_id'])
            if not diagnostico:
                return jsonify({"msg": f"Diagnóstico predefinido con ID {data['diagnostico_id']} no encontrado."}), 404
            registro.diagnostico_id = data['diagnostico_id']
            updated = True

        if 'fecha_hora' in data and data['fecha_hora']:
            try:
                new_fecha_hora = datetime.fromisoformat(data['fecha_hora'])
                if new_fecha_hora != registro.fecha_hora:
                    registro.fecha_hora = new_fecha_hora
                    updated = True
            except (ValueError, TypeError):
                return jsonify({"msg": "Formato de fecha_hora inválido. Usar formato ISO 8601 (YYYY-MM-DDTHH:MM:SS)."}), 400
        elif 'fecha_hora' in data and not data['fecha_hora']: # Permitir poner fecha_hora a null si se envía explícitamente null/vacío
             # Aunque el modelo lo define como nullable=False, si la lógica de negocio lo permite,
             # podrías cambiar el modelo o manejarlo aquí. Mantendremos nullable=False por ahora.
             return jsonify({"msg": "El campo 'fecha_hora' no puede ser nulo."}), 400

        if 'notas' in data and data['notas'] != registro.notas:
            registro.notas = data['notas'] # Puede ser null
            updated = True

        # No permitir cambiar aire_id ni registrado_por_usuario_id

        if updated:
            db.session.commit()
            return jsonify(registro.serialize()), 200
        else:
            return jsonify(registro.serialize()), 200 # O 304 Not Modified

    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"!!! ERROR SQLAlchemy al actualizar registro de diagnóstico ID {registro_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error de base de datos al actualizar el registro de diagnóstico."}), 500

    except Exception as e:
        db.session.rollback()
        print(f"!!! ERROR inesperado al actualizar registro de diagnóstico ID {registro_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor."}), 500

def buscar_o_crear_proveedor_energia(session):
    """Busca un proveedor que contenga 'energia' en su nombre o lo crea si no existe."""
    proveedor_energia = session.query(Proveedor).filter(func.lower(Proveedor.nombre).contains('energia')).first()
    if not proveedor_energia:
        print("DEBUG: Proveedor 'Energía' no encontrado, creando uno nuevo.", file=sys.stderr)
        proveedor_energia = Proveedor(nombre="Energía", email_proveedor="energia@dcce.com") # Ajusta el email si es necesario
        session.add(proveedor_energia)
        try:
            session.flush() # Para obtener el ID si es necesario antes del commit principal
            print(f"DEBUG: Proveedor 'Energía' creado con ID: {proveedor_energia.id}", file=sys.stderr)
        except Exception as e_flush:
            print(f"ERROR: Fallo al hacer flush para el nuevo proveedor Energía: {e_flush}", file=sys.stderr)
            # No relanzar aquí, dejar que el commit principal maneje el error si ocurre.
            # Si el commit falla, el proveedor no se creará.
    else:
        print(f"DEBUG: Proveedor 'Energía' encontrado con ID: {proveedor_energia.id}, Nombre: {proveedor_energia.nombre}", file=sys.stderr)
    return proveedor_energia

def crear_actividad_para_diagnostico_operatividad(session, aire_obj, registro_diagnostico, proveedor_energia_id, fecha_ocurrencia_alerta):
    """Crea una actividad para un diagnóstico de operatividad si no existe una abierta."""
    # Construir una descripción única para la actividad basada en el problema
    descripcion_problema = f"Aire: {aire_obj.nombre}, Parte: {registro_diagnostico.parte_ac.value}, Diagnóstico: {registro_diagnostico.diagnostico.nombre if registro_diagnostico.diagnostico else 'N/A'}"
    
    # Verificar si ya existe una actividad PENDIENTE o EN_PROGRESO para este problema específico
    actividad_existente = session.query(ActividadProveedor).filter(
        ActividadProveedor.proveedor_id == proveedor_energia_id,
        ActividadProveedor.descripcion.like(f"%{descripcion_problema}%"), # Buscar por la descripción del problema
        ActividadProveedor.estatus.in_([EstatusActividad.PENDIENTE, EstatusActividad.EN_PROGRESO])
    ).first()

    if not actividad_existente:
        descripcion_completa = f"Problema de operatividad detectado. {descripcion_problema}. Notas: {registro_diagnostico.notas or 'Sin notas adicionales'}."
        nueva_actividad = ActividadProveedor(
            proveedor_id=proveedor_energia_id,
            descripcion=descripcion_completa,
            fecha_ocurrencia=fecha_ocurrencia_alerta, # Fecha de la alerta/diagnóstico
            fecha_reporte=datetime.now(timezone.utc),
            estatus=EstatusActividad.PENDIENTE
        )
        session.add(nueva_actividad)
        print(f"DEBUG: Nueva actividad creada para proveedor Energía, descripción: {descripcion_completa[:100]}...", file=sys.stderr)

@api.route('/registros_diagnostico/<int:registro_id>', methods=['DELETE'])
@jwt_required()
def eliminar_registro_diagnostico_aire_route(registro_id):
    """
    Endpoint para eliminar un registro de diagnóstico histórico específico.
    Requiere autenticación (Admin/Supervisor/Tecnico).
    """
    current_user_id = get_jwt_identity()
    logged_in_user = TrackerUsuario.query.get(current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor', 'tecnico']:
        return jsonify({"msg": "Acceso no autorizado para eliminar registros de diagnóstico"}), 403

    registro = db.session.get(RegistroDiagnosticoAire, registro_id)
    if not registro:
        return jsonify({"msg": f"Registro de diagnóstico con ID {registro_id} no encontrado."}), 404

    try:
        db.session.delete(registro)
        db.session.commit()
        return '', 204 # No Content

    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"!!! ERROR SQLAlchemy al eliminar registro de diagnóstico ID {registro_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error de base de datos al eliminar el registro de diagnóstico."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"!!! ERROR inesperado al eliminar registro de diagnóstico ID {registro_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor al eliminar el registro de diagnóstico."}), 500


# --- Rutas para Lectura ---
@api.route('/aires/<int:aire_id>/lecturas', methods=['POST'])
@jwt_required() # <--- Añadido aquí
def agregar_lectura_route(aire_id):
    """
    Endpoint para agregar una nueva lectura para un aire acondicionado específico.
    Requiere autenticación.
    Recibe los datos en formato JSON.
    """
    # Opcional: Verificar si el usuario tiene permiso para agregar lecturas
    # current_user_id = get_jwt_identity()
    # logged_in_user = TrackerUsuario.query.get(current_user_id)
    # if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor', 'operador']:
    #     return jsonify({"msg": "Acceso no autorizado para agregar lecturas"}), 403

    # Verificar que el aire acondicionado existe
    aire = AireAcondicionado.query.get(aire_id)
    if not aire:
        return jsonify({"msg": f"Aire acondicionado con ID {aire_id} no encontrado."}), 404

    data = request.get_json()
    if not data:
        return jsonify({"msg": "No se recibieron datos JSON"}), 400

    # Validar campos básicos
    if 'fecha_hora' not in data or data.get('temperatura') is None: # Temperatura puede ser 0, así que None check
        return jsonify({"msg": "Faltan campos requeridos: fecha_hora, temperatura"}), 400

    es_tipo_confort = aire.tipo == 'Confort'
    humedad_data = data.get('humedad') # Puede ser None, string vacío, o un número

    # Humedad es requerida solo si el aire NO es de tipo Confort
    if not es_tipo_confort and (humedad_data is None or str(humedad_data).strip() == ''):
        return jsonify({"msg": "El campo 'humedad' es requerido para aires que no son de tipo 'Confort'."}), 400

    try:
        # Convertir fecha_hora de string a datetime
        fecha_dt = None
        try:
            # El frontend envía 'YYYY-MM-DDTHH:MM:SS'
            fecha_dt = datetime.fromisoformat(data['fecha_hora'])
        except (ValueError, TypeError):
            return jsonify({"msg": "Formato de fecha_hora inválido. Usar formato ISO 8601 (YYYY-MM-DDTHH:MM:SS)."}), 400

        # Convertir temperatura a float
        try:
            temperatura_float = float(data['temperatura'])
        except (ValueError, TypeError):
             return jsonify({"msg": "Temperatura debe ser un valor numérico."}), 400

        humedad_float = None
        if not es_tipo_confort:
            # Si no es confort, la humedad debe ser un número válido (ya validamos que no sea None/vacío arriba)
            try:
                humedad_float = float(humedad_data)
            except (ValueError, TypeError):
                return jsonify({"msg": "Humedad debe ser un valor numérico para este tipo de aire."}), 400
        elif humedad_data is not None and str(humedad_data).strip() != '':
            # Si es confort y se proporciona humedad, intentar convertirla. Si falla, se guarda como None.
            try:
                humedad_float = float(humedad_data)
            except (ValueError, TypeError):
                humedad_float = None # Opcional: podrías loggear una advertencia aquí

        # Crear nueva lectura
        nueva_lectura = Lectura(
            aire_id=aire_id,
            fecha=fecha_dt, # El modelo espera 'fecha', pero es un DateTime
            temperatura=temperatura_float,
            humedad=humedad_float
        )

        db.session.add(nueva_lectura)
        db.session.commit()

        return jsonify(nueva_lectura.serialize()), 201

    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"!!! ERROR SQLAlchemy en agregar_lectura_route para aire {aire_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error de base de datos al guardar la lectura."}), 500

    except Exception as e:
        db.session.rollback()
        print(f"!!! ERROR inesperado en agregar_lectura_route para aire {aire_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor al guardar la lectura."}), 500

@api.route('/aires/<int:aire_id>/lecturas', methods=['GET'])
@jwt_required()
def obtener_lecturas_por_aire_route(aire_id):
    """
    Endpoint para obtener todas las lecturas de un aire acondicionado específico.
    Requiere autenticación.
    """
    # Verificar que el aire acondicionado existe
    aire = AireAcondicionado.query.get(aire_id)
    if not aire:
        return jsonify({"msg": f"Aire acondicionado con ID {aire_id} no encontrado."}), 404

    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int) # Default 20 items per page
        if per_page > 100: # Limitar per_page para evitar sobrecarga
            per_page = 100

        print(f"DEBUG BACKEND: Recibida petición de lecturas para aire_id: {aire_id}, page: {page}, per_page: {per_page}", file=sys.stderr) # <-- LOG AÑADIDO

        paginated_lecturas = Lectura.query.filter_by(aire_id=aire_id)\
            .order_by(Lectura.fecha.desc())\
            .paginate(page=page, per_page=per_page, error_out=False)

        print(f"DEBUG BACKEND: Consulta de lecturas para aire {aire_id} encontró {len(paginated_lecturas.items)} items.", file=sys.stderr) # <-- LOG AÑADIDO

        lecturas_serializadas = [lectura.serialize() for lectura in paginated_lecturas.items]

        return jsonify({
            "items": lecturas_serializadas,
            "total_items": paginated_lecturas.total,
            "total_pages": paginated_lecturas.pages,
            "current_page": paginated_lecturas.page,
            "per_page": paginated_lecturas.per_page,
            "has_next": paginated_lecturas.has_next,
            "has_prev": paginated_lecturas.has_prev
        }), 200
    except Exception as e:
        print(f"!!! ERROR inesperado en obtener_lecturas_por_aire_route para aire {aire_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor al obtener lecturas."}), 500


@api.route('/diagnostico_componentes/<int:diagnostico_id>', methods=['PUT'])
@jwt_required()
def actualizar_diagnostico_componente_route(diagnostico_id):
    """
    Endpoint para actualizar un diagnóstico predefinido existente. Requiere autenticación (Admin/Supervisor).
    Recibe los datos en formato JSON.
    """
    current_user_id = get_jwt_identity()
    logged_in_user = TrackerUsuario.query.get(current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor']:
        return jsonify({"msg": "Acceso no autorizado para actualizar diagnósticos"}), 403

    diagnostico = db.session.get(DiagnosticoComponente, diagnostico_id)
    if not diagnostico:
        return jsonify({"msg": f"Diagnóstico con ID {diagnostico_id} no encontrado."}), 404

    data = request.get_json()
    if not data:
        return jsonify({"msg": "No se recibieron datos JSON"}), 400

    updated = False
    try:
        if 'nombre' in data and data['nombre'] != diagnostico.nombre:
            diagnostico.nombre = data['nombre']
            updated = True
        if 'parte_ac' in data and data['parte_ac'] != diagnostico.parte_ac.value:
            try:
                diagnostico.parte_ac = ParteACEnum(data['parte_ac'])
                updated = True
            except ValueError:
                return jsonify({"msg": f"Valor inválido para 'parte_ac': {data['parte_ac']}"}), 400
        if 'tipo_aire_sugerido' in data and data['tipo_aire_sugerido'] != diagnostico.tipo_aire_sugerido.value:
            try:
                diagnostico.tipo_aire_sugerido = TipoAireRelevanteEnum(data['tipo_aire_sugerido'])
                updated = True
            except ValueError:
                return jsonify({"msg": f"Valor inválido para 'tipo_aire_sugerido': {data['tipo_aire_sugerido']}"}), 400
        if 'descripcion_ayuda' in data and data['descripcion_ayuda'] != diagnostico.descripcion_ayuda:
            diagnostico.descripcion_ayuda = data['descripcion_ayuda']
            updated = True
        if 'activo' in data and isinstance(data['activo'], bool) and data['activo'] != diagnostico.activo:
            diagnostico.activo = data['activo']
            updated = True

        if updated:
            db.session.commit()
            return jsonify(diagnostico.serialize()), 200
        else:
            return jsonify(diagnostico.serialize()), 200 # O 304 Not Modified

    except IntegrityError:
        db.session.rollback()
        return jsonify({"msg": f"Ya existe otro diagnóstico con el nombre '{data.get('nombre')}'."}), 409
    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"!!! ERROR SQLAlchemy al actualizar diagnóstico ID {diagnostico_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error de base de datos al actualizar el diagnóstico."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"!!! ERROR inesperado al actualizar diagnóstico ID {diagnostico_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor."}), 500

@api.route('/diagnostico_componentes/<int:diagnostico_id>', methods=['DELETE'])
@jwt_required()
def eliminar_diagnostico_componente_route(diagnostico_id):
    """Elimina un diagnóstico predefinido por su ID. Requiere autenticación (Admin/Supervisor)."""
    # --- ¡IMPORTANTE: Añadir verificación de permisos! ---
    # Solo un admin o supervisor debería poder eliminar
    current_user_id = get_jwt_identity()
    logged_in_user = TrackerUsuario.query.get(current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor']:
         return jsonify({"msg": "Acceso no autorizado para eliminar diagnósticos"}), 403
    # --- Fin verificación de permisos ---

    diagnostico = db.session.get(DiagnosticoComponente, diagnostico_id)
    if not diagnostico:
        return jsonify({"msg": f"Diagnóstico con ID {diagnostico_id} no encontrado."}), 404

    try:
        db.session.delete(diagnostico)
        db.session.commit()
        return '', 204 # No Content

    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"!!! ERROR SQLAlchemy al eliminar diagnóstico ID {diagnostico_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error de base de datos al eliminar el diagnóstico."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"!!! ERROR inesperado al eliminar diagnóstico ID {diagnostico_id}: {e}", file=sys.stderr)
        traceback.print_exc()
   
@api.route('/lecturas/<int:lectura_id>', methods=['DELETE'])
@jwt_required() # <--- Añadido aquí
def eliminar_lectura_route(lectura_id):
    """
    Endpoint para eliminar una lectura específica por su ID.
    Requiere autenticación.
    """
    # --- ¡IMPORTANTE: Añadir verificación de permisos! ---
    # Solo un admin o supervisor debería poder eliminar lecturas directamente
    current_user_id = get_jwt_identity()
    logged_in_user = TrackerUsuario.query.get(current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor']:
         return jsonify({"msg": "Acceso no autorizado para eliminar lecturas"}), 403
    # --- Fin verificación de permisos ---

    lectura = Lectura.query.get(lectura_id)
    if not lectura:
        return jsonify({"msg": f"Lectura con ID {lectura_id} no encontrada."}), 404

    try:
        db.session.delete(lectura)
        db.session.commit()
        return '', 204 # No Content

    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"!!! ERROR SQLAlchemy al eliminar lectura ID {lectura_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error de base de datos al eliminar la lectura."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"!!! ERROR inesperado al eliminar lectura ID {lectura_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor al eliminar la lectura."}), 500



#Logica para cargar Excel con valores Temp humedad
def parse_date_flexible(value):
    try:
        if isinstance(value, datetime):
            return value.date()
        elif isinstance(value, time_obj):
            return None  # No es una fecha
        elif isinstance(value, str):
            # Asegúrate de probar formatos comunes de Excel, incluyendo aquellos que podrían leerse como cadenas de fecha y hora
            for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y", 
                        "%d/%m/%y", "%m/%d/%y", # año corto
                        "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", # cadenas de fecha y hora
                        "%m/%d/%Y %H:%M:%S"): 
                try:
                    return datetime.strptime(value.strip(), fmt).date()
                except ValueError:
                    continue
        # Manejar casos donde pandas podría leer fechas como Timestamps (que son objetos datetime)
        # Esto ya está cubierto por `isinstance(value, datetime)`
        return None
    except Exception: # Capturar cualquier otro error de parseo
        return None


def parse_time_flexible(value):
    try:
        if isinstance(value, time_obj):
            return value
        elif isinstance(value, datetime): # Pandas podría leer la hora como objeto datetime si la parte de la fecha es 1900-01-01
            return value.time()
        elif isinstance(value, (int, float)): # Hora de Excel como fracción de un día
            # Ejemplo: 0.25 -> 6:00, 0.5 -> 12:00
            # Esta lógica podría necesitar ajustes si se pasan números de serie de Excel directamente
            # Por ahora, asumiendo un flotante simple tipo HH.MM o similar
            if 0 <= value < 24: # Comprobación básica para flotante tipo hora
                hours = int(value)
                minutes_decimal = (value - hours) * 60
                minutes = int(minutes_decimal)
                seconds = int((minutes_decimal - minutes) * 60)
                return time_obj(hours, minutes, seconds)
        elif isinstance(value, str):
            value = value.strip()
            # Probar formatos de hora comunes
            for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p", "%I:%M:%S %p", # Formatos AM/PM
                        "%H:%M:%S.%f"): # Con microsegundos
                try:
                    return datetime.strptime(value, fmt).time()
                except ValueError:
                    continue
        return None
    except Exception: # Capturar cualquier otro error de parseo
        return None

@api.route('/lecturas/upload_excel', methods=['POST'])
@jwt_required()
def upload_lecturas_excel_route():
    current_user_id = get_jwt_identity()

    if 'file' not in request.files:
        return jsonify({"msg": "No se encontró el archivo en la solicitud."}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({"msg": "No se seleccionó ningún archivo."}), 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({"msg": "Formato de archivo no permitido. Usar .xlsx o .xls"}), 400

     # Definir contadores y errores fuera del try para que estén disponibles en el except final
    contadores = {
            'fechas_leidas': 0,
            'horas_leidas': 0,
            'aires_encontrados': 0,
            'lecturas_validas': 0,
            'lecturas_guardadas': 0
    }
    errores_detalle = []
    try:
        print("[DEBUG] Leyendo Excel con header=None para formato original del usuario.")
        df = pd.read_excel(file, header=None) # Leer sin cabeceras predefinidas por pandas
        print(f"[DEBUG] df.columns: {df.columns}")
        print(f"[DEBUG] df.index: {df.index}")
        print(f"[DEBUG] df.head():\n{df.head().to_string()}")


        if df.empty:
            return jsonify({
                "msg": "El archivo Excel está vacío."
            }), 400
        
        # Fila de fechas: índice 2 (Fila 3 en Excel)
        # Fila de horas: índice 4 (Fila 5 en Excel)
        # Fila de inicio de aires: índice 5 (Fila 6 en Excel)
        # Columna de nombres de aire: índice 1 (Columna B en Excel)
        # Columna de inicio de datos: índice 8 (Columna I en Excel)
        if df.shape[0] < 6 or df.shape[1] < 9: # Necesitamos al menos hasta la fila 6 y columna I
            return jsonify({
                "msg": "El archivo Excel no tiene suficientes filas/columnas para el formato esperado."
            }), 400

        lecturas_a_guardar = []

        # Cargar aires desde BD (nombre -> objeto AireAcondicionado)
        aires_en_db_map = {aire.nombre: aire for aire in AireAcondicionado.query.all()}
   
        print(f"[INFO] Aires en BD: {list(aires_en_db_map.keys())}")

        fechas_por_columna = {} # {col_idx: date_obj}
        horas_por_columna = {}  # {col_idx: time_obj}

        # --- Extracción de fechas (Fila 2 del Excel, índice 2 del DataFrame) ---
        fecha_actual_para_propagacion = None
        for col_idx in range(1, df.shape[1]): # Desde la columna B en adelante
            valor_fecha_celda = df.iloc[1, col_idx] # Fila 2 del Excel
            if pd.notna(valor_fecha_celda):
                fecha_parseada = parse_date_flexible(str(valor_fecha_celda).strip())
                if fecha_parseada:
                    fechas_por_columna[col_idx] = fecha_parseada
                    fecha_actual_para_propagacion = fecha_parseada
                    print(f"[DEBUG] Fecha parseada en Fila 2, Columna Excel {col_idx+1}: {fecha_parseada}")
                else: # Si no se parsea, reseteamos la propagación
                    fecha_actual_para_propagacion = None
                    print(f"[WARN] No se pudo parsear fecha en Fila 2, Columna Excel {col_idx+1}: '{valor_fecha_celda}'")
            elif fecha_actual_para_propagacion: # Si la celda está vacía pero tenemos una fecha previa
                fechas_por_columna[col_idx] = fecha_actual_para_propagacion
                print(f"[DEBUG] Fecha propagada en Fila 2, Columna Excel {col_idx+1}: {fecha_actual_para_propagacion}")

        if fechas_por_columna:
            contadores['fechas_leidas'] = len(set(fechas_por_columna.values()))
        else:
            msg_err = "No se pudieron extraer fechas válidas de la Fila 3 (columnas I en adelante)."
            errores_detalle.append(msg_err)
            print(f"[ERROR] {msg_err}")

        # --- Extracción de Horas (Fila 5 del Excel, índice 4 del DataFrame) ---
        for col_idx in range(1, df.shape[1]): # Desde la columna I en adelante
            valor_hora_celda = df.iloc[3, col_idx] # Fila 4 del Excel
            if pd.notna(valor_hora_celda):
                hora_parseada = parse_time_flexible(str(valor_hora_celda).strip())
                if hora_parseada:
                    horas_por_columna[col_idx] = hora_parseada
                    print(f"[DEBUG] Hora parseada en Fila 4, Columna Excel {col_idx+1}: {hora_parseada}")
                else:
                    print(f"[WARN] No se pudo parsear hora en Fila 4, Columna Excel {col_idx+1}: '{valor_hora_celda}'")
        
        if horas_por_columna:
            contadores['horas_leidas'] = len(horas_por_columna)
        else:
            msg_err = "No se pudieron extraer horas válidas de la Fila 4 (columnas B en adelante)."
            errores_detalle.append(msg_err)
            print(f"[ERROR] {msg_err}")

        # Si no hay fechas u horas, no podemos continuar de forma efectiva
        if not fechas_por_columna or not horas_por_columna:
            print("[ERROR] Faltan fechas u horas para procesar las lecturas.")
            return jsonify({"msg": "Faltan fechas u horas en el Excel para procesar lecturas.", "errors": errores_detalle, "summary": contadores}), 400

        # --- Procesar cada fila de datos de aires ---
        # Nombres de aires en Col B (índice 1), a partir de Fila 6 (índice 5)
        print("[DEBUG] Procesando filas de aires...")
        for fila_idx in range(5, df.shape[0]): # Desde la fila 6 del Excel en adelante
            nombre_aire_celda = str(df.iloc[fila_idx, 1]).strip() # Columna B

            # Ignorar celdas vacías, de resumen o nombres de sala que puedan estar en la columna de aires
            if nombre_aire_celda.lower() in ['nan', '', ' ', None, 'total', 'promedio', 'sala 32e', 'sala 31e', 'sala 30e']:
                continue

            aire_obj = aires_en_db_map.get(nombre_aire_celda)
            if not aire_obj:
                msg = f"Fila Excel {fila_idx+1}: Aire '{nombre_aire_celda}' no encontrado en la base de datos."
                # Evitar mensajes de error duplicados para el mismo aire no encontrado
                if not any(f"Aire '{nombre_aire_celda}' no encontrado" in err for err in errores_detalle):
                    errores_detalle.append(msg)
                print(f"[ERROR] {msg}")
                continue
            
            # Incrementar contador solo si el aire se encontró y no se ha contado antes en esta ejecución
            # (Esto es un poco más complejo de lo que parece si un aire aparece múltiples veces en el Excel,
            # pero para un conteo simple de "aires encontrados en el Excel que están en la BD" esto funciona)
            # Para un conteo único de aires encontrados, se podría usar un set.
            # Por ahora, lo contamos cada vez que se procesa una fila con un aire válido.
            contadores['aires_encontrados'] += 1 
            print(f"[INFO] Fila Excel {fila_idx+1}: Procesando Aire '{aire_obj.nombre}' (ID: {aire_obj.id}, Tipo: {aire_obj.tipo})")

            # Las lecturas para este aire están en la misma fila, desde la columna I (índice 8)
            for col_idx in range(8, df.shape[1]):
                temp_val_excel = df.iloc[fila_idx, col_idx]
                # Humedad no se lee de este formato directamente, se asume None
                hum_val_excel = None 

                fecha_para_lectura = fechas_por_columna.get(col_idx)
                hora_para_lectura = horas_por_columna.get(col_idx)

                if not fecha_para_lectura or not hora_para_lectura:
                    # Este error puede ser muy verboso si muchas columnas no tienen fecha/hora
                    # Se podría registrar una vez por columna si se desea menos verbosidad.
                    # msg = f"Fila Excel {fila_idx+1}, Columna Excel {col_idx+1}: No se pudo obtener fecha/hora para la lectura del aire '{aire_obj.nombre}'."
                    # if not any(f"Columna Excel {col_idx+1}: No se pudo obtener fecha/hora" in err for err in errores_detalle):
                    #     errores_detalle.append(msg)
                    continue

                fecha_hora_lectura = datetime.combine(fecha_para_lectura, hora_para_lectura)
                is_confort = aire_obj.tipo == 'Confort' # Asegúrate que 'Confort' sea el string exacto en tu BD

                if pd.isna(temp_val_excel) or str(temp_val_excel).strip() in ['', '#DIV/0!']:
                    continue 
                try:
                    temperatura = float(temp_val_excel)
                except (ValueError, TypeError):
                    msg = f"Fila Excel {fila_idx+1}, Columna Excel {col_idx+1}: Valor de temperatura '{temp_val_excel}' no válido para '{aire_obj.nombre}' a las {fecha_hora_lectura}. Se ignora."
                    if not any(f"Valor de temperatura '{temp_val_excel}' no válido para '{aire_obj.nombre}'" in err for err in errores_detalle):
                        errores_detalle.append(msg)
                    print(f"[WARN] {msg}")
                    continue 

                humedad = None # Asumimos que no hay humedad en este formato
                # Si el aire NO es de tipo confort, la humedad es requerida.
                # Como este formato no provee humedad, las lecturas para aires no-confort no se guardarán.
                # --- MODIFICACIÓN: Comentar o eliminar este bloque para quitar la restricción ---
                # if not is_confort:
                #     msg = f"Fila Excel {fila_idx+1}, Columna Excel {col_idx+1}: Falta humedad para '{aire_obj.nombre}' (tipo {aire_obj.tipo}) a las {fecha_hora_lectura}, y es requerido. No se puede importar esta lectura desde este formato de Excel."
                #     # Evitar spam de este error para el mismo aire
                #     if not any(f"Falta humedad para '{aire_obj.nombre}'" in err for err in errores_detalle):
                #         errores_detalle.append(msg)
                #     print(f"[WARN] {msg}")
                #     continue 

                # --- VERIFICACIÓN DE DUPLICADOS ---
                existing_lectura = Lectura.query.filter_by(aire_id=aire_obj.id, fecha=fecha_hora_lectura).first()
                if existing_lectura:
                    msg = f"Fila Excel {fila_idx+1}, Columna Excel {col_idx+1}: Ya existe una lectura para '{aire_obj.nombre}' a las {fecha_hora_lectura}. Se omite."
                    if not any(f"Ya existe una lectura para '{aire_obj.nombre}' a las {fecha_hora_lectura}" in err for err in errores_detalle): # Evitar spam
                        errores_detalle.append(msg)
                    print(f"[WARN] {msg}")
                    continue # Saltar a la siguiente lectura/columna

                contadores['lecturas_validas'] += 1
                lecturas_a_guardar.append(Lectura(
                    aire_id=aire_obj.id,
                    fecha=fecha_hora_lectura,
                    temperatura=temperatura,
                    humedad=humedad # Será None
                ))
                print(f"[INFO] Lectura válida: {aire_obj.nombre}, {fecha_hora_lectura}, T:{temperatura}, H:{humedad}")

        if lecturas_a_guardar:
            db.session.bulk_save_objects(lecturas_a_guardar)
            db.session.commit()
            contadores['lecturas_guardadas'] = len(lecturas_a_guardar)
            print("[INFO] ✅ Lecturas guardadas exitosamente en la base de datos.")
        else:
            print("[INFO] ❗ No se encontraron lecturas válidas para guardar.")

        print("\n📊 Resumen del proceso:")
        print(f"- Fechas leídas: {contadores['fechas_leidas']}")
        print(f"- Horas leídas: {contadores['horas_leidas']}")
        print(f"- Aires encontrados: {contadores['aires_encontrados']}")
        print(f"- Lecturas válidas: {contadores['lecturas_validas']}")
        print(f"- Lecturas guardadas: {contadores['lecturas_guardadas']}")
        print(f"- Errores: {len(errores_detalle)}\n")

        return jsonify({
            "msg": f"{contadores['lecturas_guardadas']} lecturas importadas de {contadores['lecturas_validas']} válidas.",
            "success_count": contadores['lecturas_guardadas'],
            "error_count": len(errores_detalle),
            "errors": errores_detalle,
            "summary": contadores
        }), 200
    except Exception as e:
        db.session.rollback()
        print(f"[CRÍTICO] Error procesando archivo Excel: {e}", file=sys.stderr)
        traceback.print_exc() 
        return jsonify({"msg": f"Error crítico al procesar el archivo Excel: {str(e)}", "errors": errores_detalle, "summary": contadores}), 500        

@api.route('/lecturas/download_excel_template', methods=['GET'])
@jwt_required() # Es buena práctica proteger también la descarga de plantillas
def download_excel_template():
    try:
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla Lecturas Temperatura"

        # Estilos
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # Azul claro
        aire_header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Verde claro

        def apply_header_style(cell):
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = header_fill
        
        # Título general (opcional)
        ws.merge_cells('A1:Q1')
        title_cell = ws['A1']
        title_cell.value = "Plantilla para Carga Masiva de Lecturas de Temperatura"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = center_align

        # --- Fila 2: FECHAS (para columnas de datos I en adelante) ---
        ws['A2'] = "FECHAS (dd/mm/yyyy) ->"
        apply_header_style(ws['A2'])
        ws.merge_cells('A2:H2')

        # --- Fila 4: HORAS (para columnas de datos I en adelante) ---
        ws['A4'] = "HORAS (HH:MM) ->"
        apply_header_style(ws['A4'])
        ws.merge_cells('A4:H4')

        # --- Fila 5: Encabezado para Nombres de Aires y Temperaturas ---
        ws['B5'] = "NOMBRE DEL AIRE (Exacto como en BD)"
        apply_header_style(ws['B5'])
        ws['B5'].fill = aire_header_fill
        ws.merge_cells('B5:H5')

        ws['I5'] = "VALORES DE TEMPERATURA (°C)"
        apply_header_style(ws['I5'])
        ws.merge_cells('I5:Q5') # Asumiendo 9 columnas de datos de ejemplo

        # Fechas y Horas de ejemplo para 9 columnas de datos (I a Q)
        today = datetime.now()
        example_dates = [
            (today).strftime("%d/%m/%Y"), (today).strftime("%d/%m/%Y"), (today).strftime("%d/%m/%Y"),
            (today + timedelta(days=1)).strftime("%d/%m/%Y"), (today + timedelta(days=1)).strftime("%d/%m/%Y"), (today + timedelta(days=1)).strftime("%d/%m/%Y"),
            (today + timedelta(days=2)).strftime("%d/%m/%Y"), (today + timedelta(days=2)).strftime("%d/%m/%Y"), (today + timedelta(days=2)).strftime("%d/%m/%Y")
        ]
        example_hours = ["06:00", "12:00", "18:00"] * 3

        for i, date_str in enumerate(example_dates):
            col_letter = chr(ord('I') + i)
            ws[f'{col_letter}2'] = date_str
            apply_header_style(ws[f'{col_letter}2'])
        
        for i, time_str in enumerate(example_hours):
            col_letter = chr(ord('I') + i)
            ws[f'{col_letter}4'] = time_str
            apply_header_style(ws[f'{col_letter}4'])

        # --- Filas de Aires Acondicionados de Ejemplo (a partir de Fila 6) ---
        example_aires = [
            "Aire Precisión Sala Principal P01",
            "Aire Confort Oficina C05",
            "UMAS Rack 17 U02"
        ]
        for row_idx, aire_name in enumerate(example_aires, start=6):
            cell_b = ws[f'B{row_idx}']
            cell_b.value = aire_name
            cell_b.border = thin_border
            cell_b.alignment = left_align
            # Dejar celdas de temperatura (I en adelante) vacías para el usuario
            for col_idx_data in range(ord('I'), ord('Q') + 1):
                 ws[f'{chr(col_idx_data)}{row_idx}'].border = thin_border

        # Ajustar anchos de columna (aproximado)
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        for col_idx_data in range(ord('I'), ord('Q') + 1):
            ws.column_dimensions[chr(col_idx_data)].width = 15

        # === Guardar el archivo ===
        wb.save(output)
        output.seek(0)

        filename = f"plantilla_carga_lecturas_temp_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"[ERROR] Error generando plantilla de Excel: {e}", file=sys.stderr)
        return jsonify({"msg": f"Error generando plantilla: {str(e)}"}), 500

@api.route('/aires/<int:aire_id>/estadisticas', methods=['GET'])
def obtener_estadisticas_por_aire_route(aire_id):
    """
    Endpoint para obtener estadísticas (promedio, min, max, desviación)
    de temperatura y humedad para un aire acondicionado específico.
    """
    # Verificar que el aire acondicionado existe
    aire = AireAcondicionado.query.get(aire_id)
    if not aire:
        return jsonify({"msg": f"Aire acondicionado con ID {aire_id} no encontrado."}), 404

    try:
        # Realizar la consulta de agregación
        result = db.session.query(
            func.avg(Lectura.temperatura).label('temp_avg'),
            func.min(Lectura.temperatura).label('temp_min'),
            func.max(Lectura.temperatura).label('temp_max'),
            func.stddev(Lectura.temperatura).label('temp_std'),
            func.avg(Lectura.humedad).label('hum_avg'),
            func.min(Lectura.humedad).label('hum_min'),
            func.max(Lectura.humedad).label('hum_max'),
            func.stddev(Lectura.humedad).label('hum_std')
        ).filter(Lectura.aire_id == aire_id).first()

        # Preparar el diccionario de respuesta, manejando valores None
        stats = {
            'temperatura_promedio': round(result.temp_avg, 2) if result.temp_avg is not None else 0,
            'temperatura_minima': round(result.temp_min, 2) if result.temp_min is not None else 0,
            'temperatura_maxima': round(result.temp_max, 2) if result.temp_max is not None else 0,
            'temperatura_desviacion': round(result.temp_std, 2) if result.temp_std is not None else 0,
            'humedad_promedio': round(result.hum_avg, 2) if result.hum_avg is not None else 0,
            'humedad_minima': round(result.hum_min, 2) if result.hum_min is not None else 0,
            'humedad_maxima': round(result.hum_max, 2) if result.hum_max is not None else 0,
            'humedad_desviacion': round(result.hum_std, 2) if result.hum_std is not None else 0,
        }

        return jsonify(stats), 200

    except Exception as e:
        print(f"!!! ERROR inesperado en obtener_estadisticas_por_aire_route para aire {aire_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        # Devolver un objeto con ceros en caso de error es más seguro para el frontend
        default_stats = {
            'temperatura_promedio': 0, 'temperatura_minima': 0, 'temperatura_maxima': 0, 'temperatura_desviacion': 0,
            'humedad_promedio': 0, 'humedad_minima': 0, 'humedad_maxima': 0, 'humedad_desviacion': 0,
        }
        return jsonify({"msg": "Error inesperado al calcular estadísticas.", "stats": default_stats}), 500

def obtener_estadisticas_generales_helper():
    """
    Helper para obtener estadísticas generales de todas las lecturas.
    Retorna un diccionario o None en caso de error.
    """
    try:
        # Usa db.session en lugar de session
        result = db.session.query(
            func.avg(Lectura.temperatura).label('temp_avg'),
            func.min(Lectura.temperatura).label('temp_min'),
            func.max(Lectura.temperatura).label('temp_max'),
            func.avg(Lectura.humedad).label('hum_avg'),
            func.min(Lectura.humedad).label('hum_min'),
            func.max(Lectura.humedad).label('hum_max'),
            func.count(Lectura.id).label('total_lecturas') # Contar todas las lecturas
        ).first()

        # Si no hay resultados o el promedio es None (sin lecturas)
        if not result or result.temp_avg is None:
            return {
                'temperatura_promedio': 0, 'temperatura_minima': 0, 'temperatura_maxima': 0,
                'humedad_promedio': 0, 'humedad_minima': 0, 'humedad_maxima': 0,
                'total_lecturas': 0
            }

        # Devolver diccionario plano con redondeo y manejo de None
        return {
            'temperatura_promedio': round(result.temp_avg, 2) if result.temp_avg is not None else 0,
            'temperatura_minima': round(result.temp_min, 2) if result.temp_min is not None else 0,
            'temperatura_maxima': round(result.temp_max, 2) if result.temp_max is not None else 0,
            'humedad_promedio': round(result.hum_avg, 2) if result.hum_avg is not None else 0,
            'humedad_minima': round(result.hum_min, 2) if result.hum_min is not None else 0,
            'humedad_maxima': round(result.hum_max, 2) if result.hum_max is not None else 0,
            'total_lecturas': result.total_lecturas if result.total_lecturas is not None else 0
        }
    except Exception as e:
        # Loggear el error es importante para depuración
        print(f"Error en obtener_estadisticas_generales_helper: {e}", file=sys.stderr)
        traceback.print_exc()
        return None # Indicar que hubo un error

def obtener_ubicaciones_helper():
    """
    Helper para obtener todas las ubicaciones únicas de los aires.
    Retorna una lista de strings o None en caso de error.
    """
    try:
        # Usa db.session
        ubicaciones_result = db.session.query(distinct(AireAcondicionado.ubicacion)).all()
        # Extraer solo el string de cada tupla, filtrando None o vacíos
        return [ubicacion[0] for ubicacion in ubicaciones_result if ubicacion[0]]
    except Exception as e:
        print(f"Error en obtener_ubicaciones_helper: {e}", file=sys.stderr)
        traceback.print_exc()
        return None # Indicar error

def obtener_estadisticas_por_ubicacion_helper(ubicacion=None):
    """
    Helper para obtener estadísticas agrupadas por ubicación.
    Retorna una lista de diccionarios o None en caso de error.
    """
    try:
        # Usa db.session
        query_aires = db.session.query(AireAcondicionado.id, AireAcondicionado.ubicacion)
        if ubicacion:
            query_aires = query_aires.filter(AireAcondicionado.ubicacion == ubicacion)

        # Agrupar IDs de aires por ubicación
        aires_por_ubicacion = {}
        all_aires = query_aires.all()
        if not all_aires:
             # Si no hay aires (en general o para esa ubicación), devuelve lista vacía
             return []

        for aire_id, loc in all_aires:
            if loc not in aires_por_ubicacion:
                aires_por_ubicacion[loc] = []
            aires_por_ubicacion[loc].append(aire_id)

        # Lista para almacenar los diccionarios de resultados directamente
        resultados = []

        # Calcular estadísticas para cada ubicación encontrada
        for loc, aires_ids in aires_por_ubicacion.items():
            if not aires_ids: continue # Seguridad extra

            # Consultar estadísticas para los aires de esta ubicación
            result = db.session.query(
                func.avg(Lectura.temperatura).label('temp_avg'),
                func.min(Lectura.temperatura).label('temp_min'),
                func.max(Lectura.temperatura).label('temp_max'),
                func.stddev(Lectura.temperatura).label('temp_std'),
                func.avg(Lectura.humedad).label('hum_avg'),
                func.min(Lectura.humedad).label('hum_min'),
                func.max(Lectura.humedad).label('hum_max'),
                func.stddev(Lectura.humedad).label('hum_std'),
                func.count(Lectura.id).label('total_lecturas') # Contar todas las lecturas
            ).filter(Lectura.aire_id.in_(aires_ids)).first()

            # Preparar diccionario de datos para esta ubicación
            stats_data = {'ubicacion': loc, 'num_aires': len(aires_ids)}

            # Añadir resultados si se encontraron lecturas
            if result and result.total_lecturas > 0:
                stats_data.update({
                    'temperatura_promedio': round(result.temp_avg, 2) if result.temp_avg is not None else 0,
                    'temperatura_min': round(result.temp_min, 2) if result.temp_min is not None else 0,
                    'temperatura_max': round(result.temp_max, 2) if result.temp_max is not None else 0,
                    'temperatura_std': round(result.temp_std, 2) if result.temp_std is not None else 0,
                    'humedad_promedio': round(result.hum_avg, 2) if result.hum_avg is not None else 0,
                    'humedad_min': round(result.hum_min, 2) if result.hum_min is not None else 0,
                    'humedad_max': round(result.hum_max, 2) if result.hum_max is not None else 0,
                    'humedad_std': round(result.hum_std, 2) if result.hum_std is not None else 0,
                    'lecturas_totales': result.total_lecturas if result.total_lecturas is not None else 0
                })
            else:
                # Añadir diccionario con ceros si no hay lecturas para los aires de esta ubicación
                stats_data.update({
                    'temperatura_promedio': 0, 'temperatura_min': 0, 'temperatura_max': 0, 'temperatura_std': 0,
                    'humedad_promedio': 0, 'humedad_min': 0, 'humedad_max': 0, 'humedad_std': 0,
                    'lecturas_totales': 0
                })
            # Añadir el diccionario a la lista de resultados
            resultados.append(stats_data)

        # Devolver la lista de diccionarios
        return resultados

    except Exception as e:
        print(f"Error en obtener_estadisticas_por_ubicacion_helper para '{ubicacion}': {e}", file=sys.stderr)
        traceback.print_exc()
        return None # Indicar error

# --- Rutas de API que usan los Helpers ---

@api.route('/estadisticas/generales', methods=['GET'])
def get_estadisticas_generales_route():
    """Endpoint para obtener estadísticas generales de todas las lecturas."""
    stats = obtener_estadisticas_generales_helper()
    if stats is None:
        # Si el helper devolvió None, hubo un error interno
        return jsonify({"msg": "Error al calcular estadísticas generales."}), 500
    # Si todo ok, devuelve el diccionario de estadísticas
    return jsonify(stats), 200

@api.route('/aires/ubicaciones', methods=['GET'])
def get_ubicaciones_route():
    """Endpoint para obtener la lista de ubicaciones únicas de los aires."""
    ubicaciones = obtener_ubicaciones_helper()
    if ubicaciones is None:
        return jsonify({"msg": "Error al obtener ubicaciones."}), 500
    # Devuelve la lista de strings de ubicaciones
    return jsonify(ubicaciones), 200

@api.route('/aires/ubicacion/<string:ubicacion>', methods=['GET'])
def get_aires_por_ubicacion_route(ubicacion):
    """Endpoint para obtener los aires acondicionados de una ubicación específica."""
    df_aires = obtener_aires_por_ubicacion_helper(ubicacion)

    if df_aires is None: # Error ocurrido en el helper
        return jsonify({"msg": f"Error al obtener aires para la ubicación '{ubicacion}'."}), 500

    # Si el DataFrame está vacío (porque no se encontraron aires), devuelve una lista vacía JSON
    if df_aires.empty:
        return jsonify([]), 200

    # Convertir DataFrame a lista de diccionarios para la respuesta JSON
    aires_list = df_aires.to_dict(orient='records')
    return jsonify(aires_list), 200

# Ruta para obtener estadísticas por ubicación (todas o una específica)
@api.route('/estadisticas/ubicacion', defaults={'ubicacion': None}, methods=['GET'])
@api.route('/estadisticas/ubicacion/<string:ubicacion>', methods=['GET'])
@jwt_required() # Asegúrate que esta ruta también esté protegida si es necesario
def get_estadisticas_por_ubicacion_route(ubicacion):
    """
    Endpoint para obtener estadísticas por ubicación.
    Si no se especifica ubicación en la URL, devuelve para todas.
    Si se especifica /estadisticas/ubicacion/NombreUbicacion, filtra por esa.
    """
    # El helper ahora devuelve una lista de diccionarios o None
    stats_list = obtener_estadisticas_por_ubicacion_helper(ubicacion)

    if stats_list is None: # Error ocurrido en el helper
        msg = f"Error al calcular estadísticas para la ubicación '{ubicacion}'." if ubicacion else "Error al calcular estadísticas por ubicación."
        return jsonify({"msg": msg}), 500

    # --- CORRECCIÓN ---
    # Verificar si la lista está vacía (en lugar de df_stats.empty)
    if not stats_list:
        return jsonify([]), 200 # Devuelve lista vacía JSON si no hay datos

    return jsonify(stats_list), 200 # Devuelve la lista directamente
# --- Helper Functions (Opcional, pero recomendado para organización) ---

def agregar_otro_equipo_helper(data):
    """
    Helper para agregar un nuevo equipo diverso.
    Retorna el objeto serializado o None en caso de error.
    Lanza excepciones específicas para manejo en la ruta.
    """
    required_fields = ['nombre', 'tipo']
    if not all(field in data for field in required_fields):
        missing = [field for field in required_fields if field not in data]
        raise ValueError(f"Faltan campos requeridos: {', '.join(missing)}") # Lanza error para 400

    # Convertir fecha si es string y es válida
    fecha_instalacion_dt = None
    fecha_str = data.get('fecha_instalacion')
    if fecha_str:
        try:
            fecha_instalacion_dt = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
             # Podrías lanzar un error aquí también si la fecha es obligatoria o debe ser válida
             print(f"Formato de fecha inválido para {fecha_str}. Se guardará como None.", file=sys.stderr)
             # raise ValueError("Formato de fecha inválido. Usar YYYY-MM-DD.")

    nuevo_equipo = OtroEquipo(
        nombre=data['nombre'],
        tipo=data['tipo'],
        ubicacion=data.get('ubicacion'),
        marca=data.get('marca'),
        modelo=data.get('modelo'),
        serial=data.get('serial'),
        codigo_inventario=data.get('codigo_inventario'),
        fecha_instalacion=fecha_instalacion_dt,
        # Asegurarse que el booleano se maneje correctamente desde JSON
        estado_operativo=bool(data.get('estado_operativo', True)),
        notas=data.get('notas')
    )
    db.session.add(nuevo_equipo)
    # El commit se maneja en la ruta para poder hacer rollback general
    return nuevo_equipo

# --- Rutas para OtroEquipo ---

@api.route('/otros_equipos', methods=['POST'])
@jwt_required() # <--- Añadido aquí
def agregar_otro_equipo_route():
    """
    Endpoint para agregar un nuevo equipo diverso (no Aire Acondicionado).
    Requiere autenticación. Recibe los datos en formato JSON.
    """
    current_user_id = get_jwt_identity()
    logged_in_user = TrackerUsuario.query.get(current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor']:
        return jsonify({"msg": "Acceso no autorizado para agregar equipos diversos"}), 403


    data = request.get_json()
    if not data:
        return jsonify({"msg": "No se recibieron datos JSON"}), 400

    try:
        # Llama al helper para crear la instancia (sin commit aún)
        nuevo_equipo = agregar_otro_equipo_helper(data)

        # Ahora intenta hacer commit
        db.session.commit()

        # Devuelve el objeto recién creado y guardado
        return jsonify(nuevo_equipo.serialize()), 201

    except ValueError as ve: # Captura errores de validación del helper
        db.session.rollback()
        return jsonify({"msg": str(ve)}), 400
    except IntegrityError as e:
        db.session.rollback()
        error_info = str(e.orig)
        msg = "Error de integridad al agregar equipo diverso."
        if 'UNIQUE constraint failed' in error_info: # Ejemplo SQLite
             if 'serial' in error_info:
                 msg = "Error: Ya existe un equipo con ese número de serie."
             elif 'codigo_inventario' in error_info:
                 msg = "Error: Ya existe un equipo con ese código de inventario."
        # Para otros DBs, el mensaje de error puede variar
        print(f"Error de integridad al agregar otro equipo: {e}", file=sys.stderr)
        return jsonify({"msg": msg}), 409 # 409 Conflict
    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"!!! ERROR SQLAlchemy en agregar_otro_equipo_route: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error de base de datos al agregar el equipo diverso."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"!!! ERROR inesperado en agregar_otro_equipo_route: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor."}), 500


@api.route('/otros_equipos', methods=['GET'])
@jwt_required() # <--- Añadido aquí
def obtener_otros_equipos_route():
    """
    Endpoint para obtener la lista de todos los equipos diversos. Requiere autenticación.
    Devuelve una lista de objetos JSON.
    """
    try:
        # Consulta directa con SQLAlchemy, ordenando por nombre
        equipos = db.session.query(OtroEquipo).order_by(OtroEquipo.nombre).all()

        # Serializar cada objeto
        equipos_serializados = [equipo.serialize() for equipo in equipos]

        return jsonify(equipos_serializados), 200

    except Exception as e:
        print(f"!!! ERROR inesperado en obtener_otros_equipos_route: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor al obtener equipos diversos."}), 500


@api.route('/otros_equipos/<int:equipo_id>', methods=['GET'])
@jwt_required() # <--- Añadido aquí
def obtener_otro_equipo_por_id_route(equipo_id):
    """
    Endpoint para obtener un equipo diverso específico por su ID. Requiere autenticación.
    """
    try:
        if equipo_id <= 0:
             return jsonify({"msg": "ID de equipo inválido."}), 400

        # Usar db.session.get() para obtener por PK
        equipo = db.session.get(OtroEquipo, equipo_id)

        if not equipo:
            return jsonify({"msg": f"Equipo diverso con ID {equipo_id} no encontrado."}), 404

        return jsonify(equipo.serialize()), 200

    except Exception as e:
        print(f"!!! ERROR inesperado en obtener_otro_equipo_por_id_route para ID {equipo_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor al obtener el equipo diverso."}), 500


@api.route('/otros_equipos/<int:equipo_id>', methods=['PUT'])
@jwt_required() # <--- Añadido aquí
def actualizar_otro_equipo_route(equipo_id):
    """
    Endpoint para actualizar un equipo diverso existente. Requiere autenticación.
    Recibe los datos en formato JSON.
    """
    # --- Opcional: Verificación de Permisos ---
    # current_user_id = get_jwt_identity()
    # logged_in_user = TrackerUsuario.query.get(current_user_id)
    # if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor']:
    #     return jsonify({"msg": "Acceso no autorizado para actualizar equipos diversos"}), 403
    # --- Fin Verificación ---

    equipo = db.session.get(OtroEquipo, equipo_id)
    if not equipo:
        return jsonify({"msg": f"Equipo diverso con ID {equipo_id} no encontrado."}), 404

    data = request.get_json()
    if not data:
        return jsonify({"msg": "No se recibieron datos JSON"}), 400

    # (Resto del código sin cambios...)
    allowed_keys = ['nombre', 'tipo', 'ubicacion', 'marca', 'modelo', 'serial',
                    'codigo_inventario', 'fecha_instalacion', 'estado_operativo', 'notas']
    updated = False

    try:
        for key, value in data.items():
            if key in allowed_keys:
                current_value = getattr(equipo, key)

                # Convertir y validar antes de asignar
                if key == 'fecha_instalacion':
                    new_date = None
                    if value: # Si se proporciona un valor
                        try:
                            new_date = datetime.strptime(value, '%Y-%m-%d').date()
                        except (ValueError, TypeError):
                            return jsonify({"msg": f"Formato de fecha inválido para {key}: {value}. Usar YYYY-MM-DD o null."}), 400
                    if new_date != current_value:
                        setattr(equipo, key, new_date)
                        updated = True
                elif key == 'estado_operativo':
                    new_bool = bool(value) # Convertir a booleano
                    if new_bool != current_value:
                        setattr(equipo, key, new_bool)
                        updated = True
                elif value != current_value: # Para otros campos (strings, etc.)
                    setattr(equipo, key, value)
                    updated = True

        if updated:
            equipo.ultima_modificacion = datetime.now(timezone.utc) # Actualizar timestamp
            db.session.commit()
            return jsonify(equipo.serialize()), 200
        else:
            # Si no hubo cambios, puedes devolver 200 OK o 304 Not Modified
            return jsonify(equipo.serialize()), 200 # O return '', 304

    except IntegrityError as e:
        db.session.rollback()
        error_info = str(e.orig)
        msg = "Error: Ya existe otro equipo con ese Serial o Código de Inventario."
        # ... (lógica similar a agregar para mensajes más específicos) ...
        print(f"Error de integridad al actualizar otro equipo {equipo_id}: {e}", file=sys.stderr)
        return jsonify({"msg": msg}), 409
    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"!!! ERROR SQLAlchemy al actualizar otro equipo ID {equipo_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error de base de datos al actualizar el equipo."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"!!! ERROR inesperado al actualizar otro equipo ID {equipo_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor."}), 500


@api.route('/otros_equipos/<int:equipo_id>', methods=['DELETE'])
@jwt_required() # <--- Añadido aquí
def eliminar_otro_equipo_route(equipo_id):
    """
    Endpoint para eliminar un equipo diverso específico por su ID. Requiere autenticación.
    Los mantenimientos asociados deberían eliminarse en cascada si la relación
    en el modelo Mantenimiento está configurada con cascade='all, delete-orphan'.
    """
    # --- ¡IMPORTANTE: Añadir verificación de permisos! ---
    # Solo un admin o supervisor debería poder eliminar
    current_user_id = get_jwt_identity()
    logged_in_user = TrackerUsuario.query.get(current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor']:
         return jsonify({"msg": "Acceso no autorizado para eliminar equipos diversos"}), 403
    # --- Fin verificación de permisos ---

    equipo = db.session.get(OtroEquipo, equipo_id)
    if not equipo:
        return jsonify({"msg": f"Equipo diverso con ID {equipo_id} no encontrado."}), 404

    try:
        # Eliminar el objeto (SQLAlchemy manejará la cascada si está configurada)
        db.session.delete(equipo)
        db.session.commit()

        return '', 204 # No Content

    except SQLAlchemyError as e: # Captura errores específicos de DB
        db.session.rollback()
        print(f"!!! ERROR SQLAlchemy al eliminar otro equipo ID {equipo_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        # Podría ser un error de restricción si la cascada no está bien o hay otras dependencias
        return jsonify({"msg": "Error de base de datos al eliminar el equipo."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"!!! ERROR inesperado al eliminar otro equipo ID {equipo_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor al eliminar el equipo."}), 500


# --- Rutas para Mantenimiento ---

# Ruta para agregar mantenimiento a un Aire Acondicionado
@api.route('/aires/<int:aire_id>/mantenimientos', methods=['POST'])
@jwt_required()
def add_mantenimiento_aire(aire_id):
    current_user_id = get_jwt_identity()
    tracker_user = TrackerUsuario.query.get(current_user_id) # Asegurarse de importar TrackerUsuario
    if not tracker_user:
        return jsonify({"msg": "Usuario no encontrado"}), 404

    aire = AireAcondicionado.query.get(aire_id)
    if not aire:
        return jsonify({"msg": "Aire Acondicionado no encontrado"}), 404

    print(f"DEBUG: Recibida petición POST para mantenimiento en aire {aire_id}", file=sys.stderr)
    print(f"DEBUG: request.form: {request.form}", file=sys.stderr)
    print(f"DEBUG: request.files: {request.files}", file=sys.stderr)


    # Start transaction
    try:
        # Extract form data
        tipo_mantenimiento = request.form.get('tipo_mantenimiento')
        descripcion = request.form.get('descripcion')
        tecnico = request.form.get('tecnico')

        # --- Get and process resolution data ---
        resolucion_alertas_data_str = request.form.get('resolucion_alertas_data')
        resolucion_alertas_data = {}
        if resolucion_alertas_data_str:
            print(f"DEBUG: resolucion_alertas_data_str recibida: {resolucion_alertas_data_str}", file=sys.stderr)
            try:
                resolucion_alertas_data = json.loads(resolucion_alertas_data_str)
                print(f"DEBUG: resolucion_alertas_data decodificada: {resolucion_alertas_data}", file=sys.stderr)
            except json.JSONDecodeError:
                print(f"ERROR: json.JSONDecodeError al decodificar resolucion_alertas_data_str", file=sys.stderr)
                db.session.rollback()
                return jsonify({"msg": "Formato de datos de resolución de alertas inválido."}), 400
            except Exception as e:
                 print(f"ERROR: Error inesperado al decodificar resolucion_alertas_data: {e}", file=sys.stderr)
                 traceback.print_exc(file=sys.stderr)
                 db.session.rollback()
                 return jsonify({"msg": "Error interno al procesar datos de resolución de alertas."}), 500
        else:
             print("DEBUG: No se recibieron datos de resolución de alertas.", file=sys.stderr)


        # Handle file upload
        imagen_file = request.files.get('imagen_file')
        imagen_datos = None
        imagen_nombre = None
        imagen_tipo = None
        if imagen_file:
            imagen_datos = imagen_file.read()
            imagen_nombre = secure_filename(imagen_file.filename)
            imagen_tipo = imagen_file.mimetype
            print(f"DEBUG: Imagen recibida: {imagen_nombre} ({imagen_tipo})", file=sys.stderr)


        # Create new Mantenimiento record
        nuevo_mantenimiento = Mantenimiento(
            aire_id=aire_id,
            fecha=datetime.now(timezone.utc),
            tipo_mantenimiento=tipo_mantenimiento,
            descripcion=descripcion,
            tecnico=tecnico,
            imagen_datos=imagen_datos,
            imagen_nombre=imagen_nombre,
            imagen_tipo=imagen_tipo,
            alertas_resueltas_info=resolucion_alertas_data_str # Guardar el string JSON original
        )
        db.session.add(nuevo_mantenimiento)
        print(f"DEBUG: Objeto Mantenimiento creado y añadido a la sesión.", file=sys.stderr)


        # --- Process resolution data and update AireAcondicionado state ---
        if resolucion_alertas_data:
            print("DEBUG: Procesando datos de resolución de alertas...", file=sys.stderr)
            for alert_key, resolucion_info in resolucion_alertas_data.items():
                print(f"DEBUG: Procesando alerta key: {alert_key}, info: {resolucion_info}", file=sys.stderr)
                # Ensure resuelta is explicitly True, not just truthy
                if resolucion_info.get('resuelta') is True:
                    print(f"DEBUG: Alerta marcada como resuelta.", file=sys.stderr)
                    componente_afectado_str = resolucion_info.get('componenteOriginal')
                    nuevo_estado_str = resolucion_info.get('nuevoEstado')
                    nuevo_diagnostico_id = resolucion_info.get('nuevoDiagnosticoId')
                    nuevas_notas = resolucion_info.get('nuevasNotas')
                    mensaje_original = resolucion_info.get('mensajeOriginal')
                    fecha_lectura_original_str = resolucion_info.get('fechaLecturaOriginal')

                    # Validate component and state
                    if componente_afectado_str and nuevo_estado_str:
                        try:
                            nuevo_estado_enum = OperativaStateEnum(nuevo_estado_str)
                            print(f"DEBUG: Componente: {componente_afectado_str}, Nuevo Estado String: '{nuevo_estado_str}', Nuevo Estado Enum: {nuevo_estado_enum}", file=sys.stderr)

                            # Update the AireAcondicionado instance
                            if componente_afectado_str == 'evaporadora':
                                print(f"DEBUG: Actualizando aire {aire_id} evaporadora_operativa de {aire.evaporadora_operativa.value} a {nuevo_estado_enum.value}", file=sys.stderr)
                                aire.evaporadora_operativa = nuevo_estado_enum
                            elif componente_afectado_str == 'condensadora':
                                print(f"DEBUG: Actualizando aire {aire_id} condensadora_operativa de {aire.condensadora_operativa.value} a {nuevo_estado_enum.value}", file=sys.stderr)
                                aire.condensadora_operativa = nuevo_estado_enum
                            else:
                                print(f"WARNING: componente_afectado_str desconocido '{componente_afectado_str}' para la clave de alerta '{alert_key}'", file=sys.stderr)
                            # Si el nuevo estado es OPERATIVA, marcar diagnósticos anteriores como solucionados
                            if nuevo_estado_enum == OperativaStateEnum.OPERATIVA and componente_afectado_str in ['evaporadora', 'condensadora']:
                                parte_enum_para_solucion = ParteACEnum(componente_afectado_str)
                                marcar_diagnosticos_como_solucionados(aire_id, parte_enum_para_solucion, db.session)

                            # --- Create new RegistroDiagnosticoAire if a new diagnosis was selected ---
                            if nuevo_diagnostico_id:
                                print(f"DEBUG: Nuevo diagnóstico ID seleccionado: {nuevo_diagnostico_id}", file=sys.stderr)
                                diagnostico_existente = DiagnosticoComponente.query.get(nuevo_diagnostico_id)
                                if diagnostico_existente:
                                    print(f"DEBUG: Diagnóstico existente encontrado: {diagnostico_existente.nombre}", file=sys.stderr)
                                    # Determine the part for the new diagnostic record
                                    parte_para_registro = ParteACEnum(componente_afectado_str) if componente_afectado_str in ['evaporadora', 'condensadora', 'general'] else diagnostico_existente.parte_ac

                                    # Attempt to parse the original alert date if available
                                    fecha_registro_diag = datetime.now(timezone.utc)
                                    if fecha_lectura_original_str:
                                        try:
                                            # Assuming ISO format from frontend
                                            fecha_registro_diag = datetime.fromisoformat(fecha_lectura_original_str.replace('Z', '+00:00'))
                                            print(f"DEBUG: Usando fecha original de alerta para diagnóstico: {fecha_registro_diag}", file=sys.stderr)
                                        except ValueError:
                                            print(f"WARNING: No se pudo parsear fechaLecturaOriginal '{fecha_lectura_original_str}' para el registro de diagnóstico. Usando fecha actual.", file=sys.stderr)
                                            # Fallback to now() if parsing fails

                                    nuevo_registro_diagnostico = RegistroDiagnosticoAire(
                                        aire_id=aire_id,
                                        parte_ac=parte_para_registro,
                                        diagnostico_id=nuevo_diagnostico_id,
                                        fecha_hora=fecha_registro_diag,
                                        notas=nuevas_notas,
                                        registrado_por_usuario_id=current_user_id,
                                    )
                                    db.session.add(nuevo_registro_diagnostico)
                                    print(f"DEBUG: Objeto RegistroDiagnosticoAire creado y añadido a la sesión para aire {aire_id}, diag_id {nuevo_diagnostico_id}", file=sys.stderr)
                                else:
                                    print(f"WARNING: DiagnosticoComponente con ID {nuevo_diagnostico_id} no encontrado. No se creará RegistroDiagnosticoAire.", file=sys.stderr)
                            else:
                                print("DEBUG: No se seleccionó un nuevo diagnóstico ID.", file=sys.stderr)


                        except ValueError as ve:
                            # Specific error for invalid Enum value string
                            print(f"ERROR: ValueError al convertir estado '{nuevo_estado_str}' a Enum para {componente_afectado_str} (clave '{alert_key}'): {ve}", file=sys.stderr)
                            # Log and continue processing other resolutions.
                        except Exception as e:
                            # Catch any other unexpected errors during processing a single resolution
                            print(f"ERROR: Error inesperado al procesar resolución para la clave de alerta '{alert_key}': {e}", file=sys.stderr)
                            traceback.print_exc(file=sys.stderr)
                            # Log and continue processing other resolutions.
                else:
                    print(f"DEBUG: Alerta key '{alert_key}' no marcada como resuelta.", file=sys.stderr)


        # --- End processing resolution data ---

        print("DEBUG: Intentando db.session.commit()...", file=sys.stderr)
        # Commit the transaction
        db.session.commit()
        print(f"DEBUG: db.session.commit() exitoso para mantenimiento en aire {aire_id}.", file=sys.stderr)


        # Refetch related data after successful commit
        # The frontend Mantenimientos.jsx already calls fetchMantenimientos, fetchAireDetails, fetchDiagnosticRecordsByAire, fetchDetailedAlerts
        # after a successful addMantenimiento, which is the correct place to trigger UI updates.

        return jsonify(nuevo_mantenimiento.serialize()), 201

    except Exception as e:
        # This catches errors that happen *after* the specific JSON/Enum errors
        # but before the commit, or during the commit itself.
        print(f"FATAL ERROR: Excepción capturada en el bloque principal para mantenimiento en aire {aire_id}: {e}", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        db.session.rollback() # Rollback changes on error
        return jsonify({"msg": "Error interno del servidor al guardar el mantenimiento."}), 500

# Ruta para agregar mantenimiento a un OtroEquipo
@api.route('/otros_equipos/<int:equipo_id>/mantenimientos', methods=['POST'])
@jwt_required()
def agregar_mantenimiento_otro_equipo_route(equipo_id):
    """
    Endpoint para agregar un registro de mantenimiento a un OtroEquipo específico.
    Requiere autenticación. Recibe datos como multipart/form-data.
    """
    # --- Opcional: Verificación de Permisos ---
    # current_user_id = get_jwt_identity()
    # logged_in_user = TrackerUsuario.query.get(current_user_id)
    # if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor', 'tecnico']:
    #     return jsonify({"msg": "Acceso no autorizado para agregar mantenimientos"}), 403
    # --- Fin Verificación ---

    # Verificar que el equipo existe
    equipo = db.session.get(OtroEquipo, equipo_id)
    if not equipo:
        return jsonify({"msg": f"Equipo diverso con ID {equipo_id} no encontrado."}), 404

    # (Resto del código sin cambios...)
    if 'tipo_mantenimiento' not in request.form or 'descripcion' not in request.form or 'tecnico' not in request.form:
        return jsonify({"msg": "Faltan campos requeridos en el formulario: tipo_mantenimiento, descripcion, tecnico"}), 400

    tipo_mantenimiento = request.form['tipo_mantenimiento']
    descripcion = request.form['descripcion']
    tecnico = request.form['tecnico']

    imagen_datos = None
    imagen_nombre = None
    imagen_tipo = None
    imagen_file = request.files.get('imagen_file')

    if imagen_file and imagen_file.filename != '':
        try:
            imagen_nombre = secure_filename(imagen_file.filename)
            imagen_tipo = imagen_file.mimetype
            imagen_datos = imagen_file.read()
        except Exception as e:
             print(f"Error leyendo archivo de imagen: {e}", file=sys.stderr)
             return jsonify({"msg": "Error al procesar el archivo de imagen."}), 400

    try:
        nuevo_mantenimiento = Mantenimiento(
            aire_id=None,
            otro_equipo_id=equipo_id,
            fecha=datetime.utcnow(),
            tipo_mantenimiento=tipo_mantenimiento,
            descripcion=descripcion,
            tecnico=tecnico,
            imagen_nombre=imagen_nombre,
            imagen_tipo=imagen_tipo,
            imagen_datos=imagen_datos
        )
        db.session.add(nuevo_mantenimiento)
        db.session.commit()
        return jsonify(nuevo_mantenimiento.serialize_with_details()), 201 # Usar serialize_with_details

    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"!!! ERROR SQLAlchemy al agregar mantenimiento para otro equipo {equipo_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error de base de datos al guardar el mantenimiento."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"!!! ERROR inesperado al agregar mantenimiento para otro equipo {equipo_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor."}), 500


# Ruta para obtener TODOS los mantenimientos (opcionalmente filtrados por query param)
@api.route('/mantenimientos', methods=['GET'])
@jwt_required()
def obtener_todos_mantenimientos_route():
    """
    Endpoint para obtener todos los registros de mantenimiento. Requiere autenticación.
    Opcionalmente filtra por ?aire_id=X o ?otro_equipo_id=Y.
    """
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        if per_page > 100: # Limitar per_page para evitar sobrecarga
            per_page = 100

        query_obj = db.session.query(Mantenimiento)

        aire_id_filter = request.args.get('aire_id', type=int)
        otro_equipo_id_filter = request.args.get('otro_equipo_id', type=int)

        if aire_id_filter:
            query_obj = query_obj.filter(Mantenimiento.aire_id == aire_id_filter)
        elif otro_equipo_id_filter:
            query_obj = query_obj.filter(Mantenimiento.otro_equipo_id == otro_equipo_id_filter)

        paginated_mantenimientos = query_obj.order_by(Mantenimiento.fecha.desc())\
            .paginate(page=page, per_page=per_page, error_out=False)

        results = [m.serialize() for m in paginated_mantenimientos.items]
        return jsonify({
            "items": results,
            "total_items": paginated_mantenimientos.total,
            "total_pages": paginated_mantenimientos.pages,
            "current_page": paginated_mantenimientos.page,
            "per_page": paginated_mantenimientos.per_page,
            "has_next": paginated_mantenimientos.has_next,
            "has_prev": paginated_mantenimientos.has_prev
        }), 200

    except Exception as e:
        print(f"!!! ERROR inesperado en obtener_todos_mantenimientos_route: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor al obtener mantenimientos."}), 500


# Ruta para obtener mantenimientos de un Aire específico
@api.route('/aires/<int:aire_id>/mantenimientos', methods=['GET'])
@jwt_required() # <--- Añadido aquí
def obtener_mantenimientos_aire_route(aire_id):
    """
    Endpoint para obtener los mantenimientos de un Aire Acondicionado específico.
    Requiere autenticación.
    """
    aire = db.session.get(AireAcondicionado, aire_id)
    if not aire:
        return jsonify({"msg": f"Aire acondicionado con ID {aire_id} no encontrado."}), 404

    try:
        mantenimientos = db.session.query(Mantenimiento)\
            .filter(Mantenimiento.aire_id == aire_id)\
            .order_by(Mantenimiento.fecha.desc())\
            .all()
        results = [m.serialize_with_details() for m in mantenimientos]
        return jsonify(results), 200

    except Exception as e:
        print(f"!!! ERROR inesperado en obtener_mantenimientos_aire_route para aire {aire_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor."}), 500


# Ruta para obtener mantenimientos de un OtroEquipo específico
@api.route('/otros_equipos/<int:equipo_id>/mantenimientos', methods=['GET'])
@jwt_required() # <--- Añadido aquí
def obtener_mantenimientos_otro_equipo_route(equipo_id):
    """
    Endpoint para obtener los mantenimientos de un OtroEquipo específico.
    Requiere autenticación.
    """
    equipo = db.session.get(OtroEquipo, equipo_id)
    if not equipo:
        return jsonify({"msg": f"Equipo diverso con ID {equipo_id} no encontrado."}), 404

    try:
        mantenimientos = db.session.query(Mantenimiento)\
            .filter(Mantenimiento.otro_equipo_id == equipo_id)\
            .order_by(Mantenimiento.fecha.desc())\
            .all()
        results = [m.serialize_with_details() for m in mantenimientos]
        return jsonify(results), 200

    except Exception as e:
        print(f"!!! ERROR inesperado en obtener_mantenimientos_otro_equipo_route para equipo {equipo_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor."}), 500


@api.route('/mantenimientos/<int:mantenimiento_id>', methods=['GET'])
@jwt_required() # <--- Añadido aquí
def obtener_mantenimiento_por_id_route(mantenimiento_id):
    """
    Endpoint para obtener un registro de mantenimiento específico por su ID.
    Requiere autenticación.
    """
    try:
        mantenimiento = db.session.get(Mantenimiento, mantenimiento_id)
        if not mantenimiento:
            return jsonify({"msg": f"Mantenimiento con ID {mantenimiento_id} no encontrado."}), 404

        if hasattr(mantenimiento, 'serialize_with_details'):
             return jsonify(mantenimiento.serialize_with_details()), 200
        else:
             return jsonify(mantenimiento.serialize()), 200

    except Exception as e:
        print(f"!!! ERROR inesperado en obtener_mantenimiento_por_id_route para ID {mantenimiento_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor."}), 500


# Ruta para obtener la imagen de un mantenimiento
@api.route('/mantenimientos/<int:mantenimiento_id>/imagen', methods=['GET'])
@jwt_required() # <--- Añadido aquí
def obtener_imagen_mantenimiento_route(mantenimiento_id):
    """
    Endpoint para obtener la imagen asociada a un mantenimiento. Requiere autenticación.
    """
    try:
        mantenimiento = db.session.get(Mantenimiento, mantenimiento_id)
        if not mantenimiento:
            return jsonify({"msg": f"Mantenimiento con ID {mantenimiento_id} no encontrado."}), 404

        if not mantenimiento.imagen_datos or not mantenimiento.imagen_tipo:
            return jsonify({"msg": "Este mantenimiento no tiene imagen asociada."}), 404

        return send_file(
            io.BytesIO(mantenimiento.imagen_datos),
            mimetype=mantenimiento.imagen_tipo,
            as_attachment=False,
            download_name=mantenimiento.imagen_nombre or f"imagen_{mantenimiento_id}"
        )

    except Exception as e:
        print(f"!!! ERROR inesperado al obtener imagen para mantenimiento ID {mantenimiento_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado al obtener la imagen."}), 500


@api.route('/mantenimientos/<int:mantenimiento_id>/imagen_base64', methods=['GET'])
@jwt_required() # <--- Añadido aquí
def obtener_imagen_mantenimiento_base64_route(mantenimiento_id):
    """
    Endpoint para obtener la imagen asociada a un mantenimiento en formato Base64.
    Requiere autenticación.
    """
    try:
        mantenimiento = db.session.get(Mantenimiento, mantenimiento_id)
        if not mantenimiento:
            return jsonify({"msg": f"Mantenimiento con ID {mantenimiento_id} no encontrado."}), 404

        if not mantenimiento.imagen_datos or not mantenimiento.imagen_tipo:
            return jsonify({"msg": "Este mantenimiento no tiene imagen asociada."}), 404

        b64_data = base64.b64encode(mantenimiento.imagen_datos).decode('utf-8')
        data_url = f"data:{mantenimiento.imagen_tipo};base64,{b64_data}"
        return jsonify({"imagen_base64": data_url}), 200

    except Exception as e:
        print(f"!!! ERROR inesperado al obtener imagen base64 para mantenimiento ID {mantenimiento_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado al obtener la imagen en base64."}), 500


@api.route('/mantenimientos/<int:mantenimiento_id>', methods=['DELETE'])
@jwt_required() # <--- Añadido aquí
def eliminar_mantenimiento_route(mantenimiento_id):
    """
    Endpoint para eliminar un registro de mantenimiento específico por su ID.
    Requiere autenticación.
    """
    # --- ¡IMPORTANTE: Añadir verificación de permisos! ---
    # Solo un admin o supervisor debería poder eliminar
    current_user_id = get_jwt_identity()
    logged_in_user = TrackerUsuario.query.get(current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor']:
         return jsonify({"msg": "Acceso no autorizado para eliminar mantenimientos"}), 403
    # --- Fin verificación de permisos ---

    mantenimiento = db.session.get(Mantenimiento, mantenimiento_id)
    if not mantenimiento:
        return jsonify({"msg": f"Mantenimiento con ID {mantenimiento_id} no encontrado."}), 404

    try:
        db.session.delete(mantenimiento)
        db.session.commit()
        return '', 204 # No Content

    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"!!! ERROR SQLAlchemy al eliminar mantenimiento ID {mantenimiento_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error de base de datos al eliminar el mantenimiento."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"!!! ERROR inesperado al eliminar mantenimiento ID {mantenimiento_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor al eliminar el mantenimiento."}), 500

# --- Rutas para UmbralConfiguracion ---

@api.route('/umbrales', methods=['POST'])
@jwt_required() # <--- Añadido aquí
def crear_umbral_configuracion_route():
    """
    Endpoint para crear una nueva configuración de umbrales. Requiere autenticación.
    Recibe los datos en formato JSON.
    """
    current_user_id = get_jwt_identity()
    logged_in_user = TrackerUsuario.query.get(current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor']:
        return jsonify({"msg": "Acceso no autorizado para crear umbrales"}), 403


    data = request.get_json()
    if not data:
        return jsonify({"msg": "No se recibieron datos JSON"}), 400

    # (Resto del código sin cambios...)
    required_fields = ['nombre', 'es_global', 'temp_min', 'temp_max', 'hum_min', 'hum_max']
    if not all(field in data for field in required_fields):
        missing = [field for field in required_fields if field not in data]
        return jsonify({"msg": f"Faltan campos requeridos: {', '.join(missing)}"}), 400

    try:
        nombre = data['nombre']
        es_global = bool(data['es_global'])
        temp_min = float(data['temp_min'])
        temp_max = float(data['temp_max'])
        hum_min = float(data['hum_min'])
        hum_max = float(data['hum_max'])
        aire_id = data.get('aire_id')
        notificar_activo = bool(data.get('notificar_activo', True))

        if temp_min >= temp_max:
            return jsonify({"msg": "temp_min debe ser menor que temp_max"}), 400
        if hum_min >= hum_max:
            return jsonify({"msg": "hum_min debe ser menor que hum_max"}), 400
        if not es_global and aire_id is None:
            return jsonify({"msg": "Se requiere aire_id si el umbral no es global (es_global=false)"}), 400
        if es_global:
            aire_id = None

        if aire_id is not None:
            aire = db.session.get(AireAcondicionado, aire_id)
            if not aire:
                 return jsonify({"msg": f"Aire acondicionado con ID {aire_id} no encontrado."}), 404

        nuevo_umbral = UmbralConfiguracion(
            nombre=nombre,
            es_global=es_global,
            aire_id=aire_id,
            temp_min=temp_min,
            temp_max=temp_max,
            hum_min=hum_min,
            hum_max=hum_max,
            notificar_activo=notificar_activo,
        )

        db.session.add(nuevo_umbral)
        db.session.commit()

        return jsonify(nuevo_umbral.serialize()), 201

    except (ValueError, TypeError) as ve:
        db.session.rollback()
        return jsonify({"msg": f"Error en el tipo de dato: {ve}. Asegúrate que los umbrales sean números."}), 400
    except IntegrityError as e:
        db.session.rollback()
        print(f"Error de integridad al crear umbral: {e}", file=sys.stderr)
        return jsonify({"msg": "Error de integridad al crear el umbral."}), 409
    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"!!! ERROR SQLAlchemy en crear_umbral_configuracion_route: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error de base de datos al crear el umbral."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"!!! ERROR inesperado en crear_umbral_configuracion_route: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor."}), 500


@api.route('/umbrales', methods=['GET'])
@jwt_required() # <--- Añadido aquí
def obtener_umbrales_configuracion_route():
    """
    Endpoint para obtener las configuraciones de umbrales. Requiere autenticación.
    Filtra por ?aire_id=X o ?solo_globales=true.
    Si no hay filtros, devuelve todos (globales y específicos).
    """
    try:
        query = db.session.query(UmbralConfiguracion)

        aire_id_filter = request.args.get('aire_id', type=int)
        solo_globales_filter = request.args.get('solo_globales', 'false').lower() == 'true'

        if solo_globales_filter:
            query = query.filter(UmbralConfiguracion.es_global == True)
        elif aire_id_filter:
            query = query.filter(
                or_(
                    UmbralConfiguracion.aire_id == aire_id_filter,
                    UmbralConfiguracion.es_global == True
                )
            )

        umbrales = query.order_by(UmbralConfiguracion.es_global.desc(), UmbralConfiguracion.nombre).all()
        results = [u.serialize() for u in umbrales]

        return jsonify(results), 200

    except Exception as e:
        print(f"!!! ERROR inesperado en obtener_umbrales_configuracion_route: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor al obtener umbrales."}), 500


@api.route('/umbrales/<int:umbral_id>', methods=['GET'])
@jwt_required() # <--- Añadido aquí
def obtener_umbral_por_id_route(umbral_id):
    """
    Endpoint para obtener una configuración de umbral específica por su ID.
    Requiere autenticación.
    """
    try:
        umbral = db.session.get(UmbralConfiguracion, umbral_id)
        if not umbral:
            return jsonify({"msg": f"Umbral con ID {umbral_id} no encontrado."}), 404

        return jsonify(umbral.serialize()), 200

    except Exception as e:
        print(f"!!! ERROR inesperado en obtener_umbral_por_id_route para ID {umbral_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor."}), 500


@api.route('/umbrales/<int:umbral_id>', methods=['PUT'])
@jwt_required() 
def actualizar_umbral_configuracion_route(umbral_id):
    """
    Endpoint para actualizar una configuración de umbral existente. Requiere autenticación.
    No permite cambiar 'es_global' ni 'aire_id'.
    """
    
    current_user_id = get_jwt_identity()
    logged_in_user = TrackerUsuario.query.get(current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor']:
        return jsonify({"msg": "Acceso no autorizado para actualizar umbrales"}), 403
    
    umbral = db.session.get(UmbralConfiguracion, umbral_id)
    if not umbral:
        return jsonify({"msg": f"Umbral con ID {umbral_id} no encontrado."}), 404

    data = request.get_json()
    if not data:
        return jsonify({"msg": "No se recibieron datos JSON"}), 400

    # (Resto del código sin cambios...)
    allowed_updates = ['nombre', 'temp_min', 'temp_max', 'hum_min', 'hum_max', 'notificar_activo']
    updated = False

    try:
        new_temp_min = float(data.get('temp_min', umbral.temp_min))
        new_temp_max = float(data.get('temp_max', umbral.temp_max))
        new_hum_min = float(data.get('hum_min', umbral.hum_min))
        new_hum_max = float(data.get('hum_max', umbral.hum_max))

        if new_temp_min >= new_temp_max:
            return jsonify({"msg": "temp_min debe ser menor que temp_max"}), 400
        if new_hum_min >= new_hum_max:
            return jsonify({"msg": "hum_min debe ser menor que hum_max"}), 400

        if 'nombre' in data and data['nombre'] != umbral.nombre:
            umbral.nombre = data['nombre']
            updated = True
        if new_temp_min != umbral.temp_min:
            umbral.temp_min = new_temp_min
            updated = True
        if new_temp_max != umbral.temp_max:
            umbral.temp_max = new_temp_max
            updated = True
        if new_hum_min != umbral.hum_min:
            umbral.hum_min = new_hum_min
            updated = True
        if new_hum_max != umbral.hum_max:
            umbral.hum_max = new_hum_max
            updated = True
        if 'notificar_activo' in data:
             new_notify = bool(data['notificar_activo'])
             if new_notify != umbral.notificar_activo:
                 umbral.notificar_activo = new_notify
                 updated = True

        if updated:
            db.session.commit()
            return jsonify(umbral.serialize()), 200
        else:
            return jsonify(umbral.serialize()), 200

    except (ValueError, TypeError) as ve:
        db.session.rollback()
        return jsonify({"msg": f"Error en el tipo de dato: {ve}. Asegúrate que los umbrales sean números."}), 400
    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"!!! ERROR SQLAlchemy al actualizar umbral ID {umbral_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error de base de datos al actualizar el umbral."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"!!! ERROR inesperado al actualizar umbral ID {umbral_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor."}), 500


@api.route('/umbrales/<int:umbral_id>', methods=['DELETE'])
@jwt_required() # <--- Añadido aquí
def eliminar_umbral_configuracion_route(umbral_id):
    """
    Endpoint para eliminar una configuración de umbral por su ID. Requiere autenticación.
    """
    current_user_id = get_jwt_identity()
    logged_in_user = TrackerUsuario.query.get(current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor']:
         return jsonify({"msg": "Acceso no autorizado para eliminar umbrales"}), 403

    umbral = db.session.get(UmbralConfiguracion, umbral_id)
    if not umbral:
        return jsonify({"msg": f"Umbral con ID {umbral_id} no encontrado."}), 404

    try:
        db.session.delete(umbral)
        db.session.commit()
        return '', 204 # No Content

    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"!!! ERROR SQLAlchemy al eliminar umbral ID {umbral_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error de base de datos al eliminar el umbral."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"!!! ERROR inesperado al eliminar umbral ID {umbral_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor al eliminar el umbral."}), 500

def find_energia_provider():
    """
    Busca un proveedor con el nombre 'Energía'.
    Retorna el objeto Proveedor o None si no se encuentra.
    """
    try:
        # Usar ilike para búsqueda insensible a mayúsculas/minúsculas si es necesario
        energia_provider = db.session.query(Proveedor).filter(func.lower(Proveedor.nombre) == 'energía').first()
        if not energia_provider:
             energia_provider = db.session.query(Proveedor).filter(func.lower(Proveedor.nombre) == 'energia').first() # Sin tilde
        return energia_provider
    except Exception as e:
        print(f"Error buscando proveedor 'Energía': {e}", file=sys.stderr)
        return None

def create_operatividad_activity_if_needed(alert_data, proveedor_energia_id):
    """
    Crea una actividad para el proveedor 'Energía' si una alerta de operatividad
    no tiene una actividad 'Pendiente' o 'En Progreso' asociada.
    """
    try:
        # Verificar si ya existe una actividad abierta para esta alerta específica
        # Usaremos el mensaje de la alerta para identificarla (podría mejorarse con un ID de alerta único si se implementa)
        existing_activity = db.session.query(ActividadProveedor).filter(
            ActividadProveedor.proveedor_id == proveedor_energia_id,
            ActividadProveedor.descripcion.like(f"%{alert_data['mensaje']}%Aire: {alert_data['aire_nombre']}%"), # Buscar por mensaje y nombre del aire
            ActividadProveedor.estatus.in_([EstatusActividad.PENDIENTE, EstatusActividad.EN_PROGRESO])
        ).first()

        if not existing_activity:
            descripcion_actividad = (
                f"Alerta de Operatividad: {alert_data['mensaje']} "
                f"Aire: {alert_data['aire_nombre']} (ID: {alert_data['aire_id']}) "
                f"Ubicación: {alert_data['aire_ubicacion']}. "
                f"Estado detectado: {alert_data['valor_actual']}."
            )
            nueva_actividad = ActividadProveedor(
                proveedor_id=proveedor_energia_id,
                descripcion=descripcion_actividad,
                fecha_ocurrencia=datetime.fromisoformat(alert_data['fecha_lectura']), # Usar la fecha de la alerta
                fecha_reporte=datetime.now(timezone.utc), # Fecha actual del reporte
                estatus=EstatusActividad.PENDIENTE
            )
            db.session.add(nueva_actividad)
            db.session.commit()
            print(f"Actividad de operatividad creada para proveedor Energía, aire {alert_data['aire_nombre']}.", file=sys.stderr)
    except Exception as e:
        db.session.rollback()
        print(f"Error creando actividad de operatividad para proveedor Energía: {e}", file=sys.stderr)

def obtener_detalles_alertas_activas_helper():
    alertas_detalladas = []
    now_iso = datetime.now(timezone.utc).isoformat()

    # 1. Obtener todos los aires acondicionados con su información básica
    aires = AireAcondicionado.query.all()
    if not aires:
        return []
    # Buscar proveedor "Energía" una vez
    energia_provider = find_energia_provider()

    # Pre-fetch latest diagnostic records for all aires and relevant parts
    all_latest_diagnostics_subquery = db.session.query(
        RegistroDiagnosticoAire.aire_id,
        RegistroDiagnosticoAire.parte_ac,
        func.max(RegistroDiagnosticoAire.fecha_hora).label('max_fecha_diag')
    ).group_by(RegistroDiagnosticoAire.aire_id, RegistroDiagnosticoAire.parte_ac).subquery()

    all_latest_diagnostics_records = db.session.query(RegistroDiagnosticoAire).join(
        all_latest_diagnostics_subquery,
        (RegistroDiagnosticoAire.aire_id == all_latest_diagnostics_subquery.c.aire_id) &
        (RegistroDiagnosticoAire.parte_ac == all_latest_diagnostics_subquery.c.parte_ac) &
        (RegistroDiagnosticoAire.fecha_hora == all_latest_diagnostics_subquery.c.max_fecha_diag)
    ).options(db.joinedload(RegistroDiagnosticoAire.diagnostico)).all() # Eager load diagnostico

    latest_diagnostics_map = {} # {(aire_id, parte_ac_enum): {"nombre": "...", "notas": "..."}}
    for record in all_latest_diagnostics_records:
        if record.diagnostico: # Ensure diagnostico relationship is loaded and exists
            latest_diagnostics_map[(record.aire_id, record.parte_ac)] = { # Use the Enum member as key
                "nombre": record.diagnostico.nombre,
                "notas": record.notas
            }

    aires_map = {aire.id: aire for aire in aires}

    # 2. Verificar estado operativo para todos los aires
    for aire_id, aire_obj in aires_map.items(): # Iterar sobre el mapa para acceso rápido
        if aire_obj.evaporadora_operativa == OperativaStateEnum.NO_OPERATIVA:
            alert_data = {
                "aire_id": aire_obj.id, "aire_nombre": aire_obj.nombre, "aire_ubicacion": aire_obj.ubicacion,
                "alerta_tipo": "Operatividad", "mensaje": "Evaporadora no operativa.", # Mensaje principal
                "valor_actual": OperativaStateEnum.NO_OPERATIVA.value, # Estado del componente
                "limite_violado": "Debe estar Operativa", "fecha_lectura": now_iso, # Fecha de detección del estado
                "diagnostico_nombre": latest_diagnostics_map.get((aire_obj.id, ParteACEnum.EVAPORADORA), {}).get("nombre"),
                "diagnostico_notas": latest_diagnostics_map.get((aire_obj.id, ParteACEnum.EVAPORADORA), {}).get("notas"),
                "requiere_proveedor_energia": energia_provider is None # Nueva bandera
            }
            alertas_detalladas.append(alert_data)
            if energia_provider:
                create_operatividad_activity_if_needed(alert_data, energia_provider.id)
            # else:
                # Si no hay proveedor energía, la bandera 'requiere_proveedor_energia' lo indicará al frontend
        elif aire_obj.evaporadora_operativa == OperativaStateEnum.PARCIALMENTE_OPERATIVA:
            alert_data = { # Definir alert_data también aquí
                "aire_id": aire_obj.id, "aire_nombre": aire_obj.nombre, "aire_ubicacion": aire_obj.ubicacion,
                "alerta_tipo": "Operatividad", "mensaje": "Evaporadora parcialmente operativa.",
                "valor_actual": OperativaStateEnum.PARCIALMENTE_OPERATIVA.value,
                "limite_violado": "Debe estar Operativa", "fecha_lectura": now_iso,
                "diagnostico_nombre": latest_diagnostics_map.get((aire_obj.id, ParteACEnum.EVAPORADORA), {}).get("nombre"),
                "diagnostico_notas": latest_diagnostics_map.get((aire_obj.id, ParteACEnum.EVAPORADORA), {}).get("notas"),
                "requiere_proveedor_energia": energia_provider is None # Nueva bandera
            }
            alertas_detalladas.append(alert_data)
            if energia_provider:
                create_operatividad_activity_if_needed(alert_data, energia_provider.id)
            # else:
                # Si no hay proveedor energía, la bandera 'requiere_proveedor_energia' lo indicará al frontend


        if aire_obj.condensadora_operativa == OperativaStateEnum.NO_OPERATIVA:
            alert_data = { # Definir alert_data también aquí
                "aire_id": aire_obj.id, "aire_nombre": aire_obj.nombre, "aire_ubicacion": aire_obj.ubicacion,
                "alerta_tipo": "Operatividad", "mensaje": "Condensadora no operativa.",
                "valor_actual": OperativaStateEnum.NO_OPERATIVA.value,
                "limite_violado": "Debe estar Operativa", "fecha_lectura": now_iso,
                "diagnostico_nombre": latest_diagnostics_map.get((aire_obj.id, ParteACEnum.CONDENSADORA), {}).get("nombre"),
                "diagnostico_notas": latest_diagnostics_map.get((aire_obj.id, ParteACEnum.CONDENSADORA), {}).get("notas"),
                "requiere_proveedor_energia": energia_provider is None # Nueva bandera
            }
            alertas_detalladas.append(alert_data)
            if energia_provider:
                create_operatividad_activity_if_needed(alert_data, energia_provider.id)
            # else:
                # Si no hay proveedor energía, la bandera 'requiere_proveedor_energia' lo indicará al frontend
        elif aire_obj.condensadora_operativa == OperativaStateEnum.PARCIALMENTE_OPERATIVA:
            alert_data = { # Definir alert_data también aquí
                "aire_id": aire_obj.id, "aire_nombre": aire_obj.nombre, "aire_ubicacion": aire_obj.ubicacion,
                "alerta_tipo": "Operatividad", "mensaje": "Condensadora parcialmente operativa.",
                "valor_actual": OperativaStateEnum.PARCIALMENTE_OPERATIVA.value,
                "limite_violado": "Debe estar Operativa", "fecha_lectura": now_iso,
                "diagnostico_nombre": latest_diagnostics_map.get((aire_obj.id, ParteACEnum.CONDENSADORA), {}).get("nombre"),
                "diagnostico_notas": latest_diagnostics_map.get((aire_obj.id, ParteACEnum.CONDENSADORA), {}).get("notas"),
                "requiere_proveedor_energia": energia_provider is None # Nueva bandera
            }
            alertas_detalladas.append(alert_data)
            if energia_provider:
                create_operatividad_activity_if_needed(alert_data, energia_provider.id)
            # else:
                # Si no hay proveedor energía, la bandera 'requiere_proveedor_energia' lo indicará al frontend


    # 3. Obtener la última lectura de cada aire
    subquery_ultimas_fechas = db.session.query(
        Lectura.aire_id,
        func.max(Lectura.fecha).label('max_fecha')
    ).group_by(Lectura.aire_id).subquery()

    ultimas_lecturas_records = db.session.query(Lectura).join(
        subquery_ultimas_fechas, # Unirse a la subconsulta para obtener solo las últimas lecturas
        (Lectura.aire_id == subquery_ultimas_fechas.c.aire_id) & (Lectura.fecha == subquery_ultimas_fechas.c.max_fecha)
    ).all()

    # 4. Pre-cargar umbrales activos
    global_umbrales = UmbralConfiguracion.query.filter(
        UmbralConfiguracion.notificar_activo == True,
        UmbralConfiguracion.es_global == True
    ).all()
    specific_umbrales_raw = UmbralConfiguracion.query.filter(
        UmbralConfiguracion.notificar_activo == True,
        UmbralConfiguracion.es_global == False,
        UmbralConfiguracion.aire_id != None
    ).all()
    specific_umbrales_dict = {}
    for u_spec in specific_umbrales_raw:
        if u_spec.aire_id not in specific_umbrales_dict:
            specific_umbrales_dict[u_spec.aire_id] = []
        specific_umbrales_dict[u_spec.aire_id].append(u_spec)

    # 5. Verificar cada última lectura contra los umbrales aplicables
    for lectura in ultimas_lecturas_records:
        aire_obj = aires_map.get(lectura.aire_id)
        if not aire_obj:
             continue
        # Solo alertas ambientales para aires que no estén completamente NO_OPERATIVOS en AMBAS unidades.
        # Si al menos una unidad está OPERATIVA o PARCIALMENTE_OPERATIVA, se verifican umbrales.
        if aire_obj.evaporadora_operativa == OperativaStateEnum.NO_OPERATIVA or aire_obj.condensadora_operativa == OperativaStateEnum.NO_OPERATIVA:
            continue # Solo alertas ambientales para aires operativos

        umbrales_specific_para_aire = specific_umbrales_dict.get(lectura.aire_id, [])
        umbrales_aplicables = global_umbrales + umbrales_specific_para_aire

        if not umbrales_aplicables:
            continue

        for umbral in umbrales_aplicables:
            try:
                temp_lectura = float(lectura.temperatura)
                hum_lectura = float(lectura.humedad) if lectura.humedad is not None else None
                temp_min, temp_max = float(umbral.temp_min), float(umbral.temp_max)
                hum_min, hum_max = float(umbral.hum_min), float(umbral.hum_max)

                if temp_lectura < temp_min:
                    alertas_detalladas.append({
                        "aire_id": aire_obj.id, "aire_nombre": aire_obj.nombre, "aire_ubicacion": aire_obj.ubicacion,
                        "alerta_tipo": "Temperatura Baja", "mensaje": f"Umbral: '{umbral.nombre}'.",
                        "valor_actual": f"{temp_lectura:.1f}°C", "limite_violado": f"Min: {temp_min:.1f}°C", "fecha_lectura": lectura.fecha.isoformat()
                    })
                elif temp_lectura > temp_max:
                    alertas_detalladas.append({
                        "aire_id": aire_obj.id, "aire_nombre": aire_obj.nombre, "aire_ubicacion": aire_obj.ubicacion,
                        "alerta_tipo": "Temperatura Alta", "mensaje": f"Umbral: '{umbral.nombre}'.",
                        "valor_actual": f"{temp_lectura:.1f}°C", "limite_violado": f"Max: {temp_max:.1f}°C", "fecha_lectura": lectura.fecha.isoformat()
                    })
                if hum_lectura is not None:
                    if hum_lectura < hum_min:
                        alertas_detalladas.append({"aire_id": aire_obj.id, "aire_nombre": aire_obj.nombre, "aire_ubicacion": aire_obj.ubicacion,"alerta_tipo": "Humedad Baja","mensaje": f"Umbral: '{umbral.nombre}'.","valor_actual": f"{hum_lectura:.1f}%", "limite_violado": f"Min: {hum_min:.1f}%", "fecha_lectura": lectura.fecha.isoformat()})
                    elif hum_lectura > hum_max:
                        alertas_detalladas.append({"aire_id": aire_obj.id, "aire_nombre": aire_obj.nombre, "aire_ubicacion": aire_obj.ubicacion,"alerta_tipo": "Humedad Alta","mensaje": f"Umbral: '{umbral.nombre}'.","valor_actual": f"{hum_lectura:.1f}%", "limite_violado": f"Max: {hum_max:.1f}%", "fecha_lectura": lectura.fecha.isoformat()})
            except Exception as e_compare:
                print(f"Error comparando lectura para aire {aire_obj.id} con umbral {umbral.id}: {e_compare}", file=sys.stderr)
    return alertas_detalladas

def verificar_lectura_dentro_umbrales_helper(aire_id, temperatura, humedad):
    """
    Helper para verificar si una lectura está dentro de los umbrales configurados.

    Args:
        aire_id (int): ID del aire acondicionado.
        temperatura (float): Temperatura a verificar.
        humedad (float): Humedad a verificar.

    Returns:
        dict: {'dentro_limite': bool, 'alertas': list} o None en caso de error grave.
    """
    try:
        # Obtener umbrales aplicables (específicos del aire + globales)
        # Directamente con SQLAlchemy, sin Pandas
        umbrales_aplicables = db.session.query(UmbralConfiguracion).filter(
            UmbralConfiguracion.notificar_activo == True, # Solo umbrales activos
            or_(
                UmbralConfiguracion.aire_id == aire_id,
                UmbralConfiguracion.es_global == True
            )
        ).all()

        if not umbrales_aplicables:
            # Si no hay umbrales activos configurados, está dentro de límites
            return {
                'dentro_limite': True,
                'alertas': []
            }

        alertas = []

        # Verificar cada umbral aplicable
        for umbral in umbrales_aplicables:
            # Verificar temperatura
            if temperatura < umbral.temp_min:
                alertas.append({
                    'tipo': 'temperatura',
                    'umbral_id': umbral.id,
                    'umbral_nombre': umbral.nombre,
                    'valor': temperatura,
                    'limite': umbral.temp_min,
                    'mensaje': f"Temperatura ({temperatura}°C) por debajo del mínimo ({umbral.temp_min}°C) - Umbral '{umbral.nombre}'"
                })
            elif temperatura > umbral.temp_max:
                alertas.append({
                    'tipo': 'temperatura',
                    'umbral_id': umbral.id,
                    'umbral_nombre': umbral.nombre,
                    'valor': temperatura,
                    'limite': umbral.temp_max,
                    'mensaje': f"Temperatura ({temperatura}°C) por encima del máximo ({umbral.temp_max}°C) - Umbral '{umbral.nombre}'"
                })

            # Verificar humedad
            if humedad < umbral.hum_min:
                alertas.append({
                    'tipo': 'humedad',
                    'umbral_id': umbral.id,
                    'umbral_nombre': umbral.nombre,
                    'valor': humedad,
                    'limite': umbral.hum_min,
                    'mensaje': f"Humedad ({humedad}%) por debajo del mínimo ({umbral.hum_min}%) - Umbral '{umbral.nombre}'"
                })
            elif humedad > umbral.hum_max:
                alertas.append({
                    'tipo': 'humedad',
                    'umbral_id': umbral.id,
                    'umbral_nombre': umbral.nombre,
                    'valor': humedad,
                    'limite': umbral.hum_max,
                    'mensaje': f"Humedad ({humedad}%) por encima del máximo ({umbral.hum_max}%) - Umbral '{umbral.nombre}'"
                })

        # Devolver resultado
        return {
            'dentro_limite': len(alertas) == 0,
            'alertas': alertas
        }

    except Exception as e:
        print(f"Error en verificar_lectura_dentro_umbrales_helper para aire {aire_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return None # Indicar error


# --- Ruta de API para Verificar Umbrales ---

@api.route('/aires/<int:aire_id>/verificar_umbrales', methods=['GET'])
@jwt_required()
def verificar_umbrales_route(aire_id):
    """
    Endpoint para verificar si una lectura de temperatura y humedad
    está dentro de los umbrales configurados para un aire específico. Requiere autenticación.
    Recibe 'temp' y 'hum' como query parameters.
    Ej: /aires/1/verificar_umbrales?temp=28.5&hum=65
    """
    # Verificar que el aire acondicionado existe
    aire = db.session.get(AireAcondicionado, aire_id)
    if not aire:
        return jsonify({"msg": f"Aire acondicionado con ID {aire_id} no encontrado."}), 404

    # Obtener parámetros de la query string
    temp_str = request.args.get('temp')
    hum_str = request.args.get('hum')

    if temp_str is None or hum_str is None:
        return jsonify({"msg": "Parámetros 'temp' y 'hum' son requeridos en la URL."}), 400

    try:
        temperatura = float(temp_str)
        humedad = float(hum_str)
    except (ValueError, TypeError):
        return jsonify({"msg": "'temp' y 'hum' deben ser valores numéricos."}), 400

    # Llamar al helper para hacer la verificación
    resultado = verificar_lectura_dentro_umbrales_helper(aire_id, temperatura, humedad)

    if resultado is None:
        return jsonify({"msg": "Error al verificar los umbrales."}), 500

    return jsonify(resultado), 200

def contar_entidades_helper():
    """Helper para contar diferentes tipos de entidades."""
    counts = {}
    try:
        counts['aires'] = db.session.query(func.count(AireAcondicionado.id)).scalar() or 0
    except Exception as e:
        print(f"Error al contar aires: {e}", file=sys.stderr)
        counts['aires'] = -1 # Indicar error

    try:
        counts['lecturas'] = db.session.query(func.count(Lectura.id)).scalar() or 0
    except Exception as e:
        print(f"Error al contar lecturas: {e}", file=sys.stderr)
        counts['lecturas'] = -1

    try:
        counts['mantenimientos'] = db.session.query(func.count(Mantenimiento.id)).scalar() or 0
    except Exception as e:
        print(f"Error al contar mantenimientos: {e}", file=sys.stderr)
        counts['mantenimientos'] = -1

    try:
        counts['otros_equipos'] = db.session.query(func.count(OtroEquipo.id)).scalar() or 0
    except Exception as e:
        print(f"Error al contar otros equipos: {e}", file=sys.stderr)
        counts['otros_equipos'] = -1

    return counts

def contar_alertas_activas_helper():
    """
    Helper para contar el número de aires con al menos una alerta activa
    basada en su última lectura y los umbrales activos. (OPTIMIZADO)
    """
    print("\n--- [DEBUG] Iniciando contar_alertas_activas_helper ---")
    try:
        aires_con_alerta = set() # Usar un set para no contar dos veces el mismo aire

        # --- 1. Contar Alertas de Operatividad ---
        print("[DEBUG] Contando alertas de operatividad...")
        aires_con_problemas_operatividad = db.session.query(AireAcondicionado.id).filter(
            or_(
                AireAcondicionado.evaporadora_operativa != OperativaStateEnum.OPERATIVA,
                AireAcondicionado.condensadora_operativa != OperativaStateEnum.OPERATIVA
            )
        ).all()

        for aire_tuple in aires_con_problemas_operatividad:
            aires_con_alerta.add(aire_tuple[0]) # Añadir el ID del aire
        
        print(f"[DEBUG] Aires con problemas de operatividad: {len(aires_con_problemas_operatividad)} (IDs: {aires_con_alerta})")

        # --- 2. Contar Alertas Ambientales (basadas en umbrales) ---
        print("[DEBUG] Contando alertas ambientales...")
        # 2.1. Obtener la última lectura de cada aire
        subquery = db.session.query(
            Lectura.aire_id,
            func.max(Lectura.fecha).label('max_fecha')
        ).group_by(Lectura.aire_id).subquery()

        ultimas_lecturas = db.session.query(Lectura).join(
            subquery,
            (Lectura.aire_id == subquery.c.aire_id) & (Lectura.fecha == subquery.c.max_fecha)
        ).all()
        print(f"[DEBUG] Últimas lecturas encontradas para alertas ambientales: {len(ultimas_lecturas)}")

        # 2.2. Pre-cargar y organizar umbrales
        # Pre-fetch global thresholds once
        global_umbrales = db.session.query(UmbralConfiguracion).filter(
            UmbralConfiguracion.notificar_activo == True,
            UmbralConfiguracion.es_global == True
        ).all()
        
        # Pre-fetch specific thresholds into a dictionary {aire_id: [umbral1, umbral2]}
        specific_umbrales_raw = db.session.query(UmbralConfiguracion).filter(
            UmbralConfiguracion.notificar_activo == True,
            UmbralConfiguracion.es_global == False,
            UmbralConfiguracion.aire_id != None # Asegura que aire_id no sea null
        ).all()
        specific_umbrales_dict = {}
        for u in specific_umbrales_raw:
            if u.aire_id not in specific_umbrales_dict:
                specific_umbrales_dict[u.aire_id] = []
            specific_umbrales_dict[u.aire_id].append(u)
        
        print(f"[DEBUG] Global umbrales ACTIVOS: {len(global_umbrales)}, Specific umbrales ACTIVOS: {len(specific_umbrales_raw)}")

        # Si no hay ningún umbral activo (ni global ni específico), no puede haber alertas ambientales
        if not global_umbrales and not specific_umbrales_dict:
             print("[DEBUG] No hay umbrales activos para alertas ambientales. Saltando esta parte.")
        else:
            # 2.3. Verificar cada última lectura contra los umbrales aplicables
            print("[DEBUG] Iniciando verificación de lecturas vs umbrales para alertas ambientales...")
            for lectura in ultimas_lecturas:
                aire_actual = db.session.get(AireAcondicionado, lectura.aire_id)
                if not aire_actual: 
                    print(f"[WARN-DEBUG] Aire ID: {lectura.aire_id} de la lectura no encontrado en la tabla de aires. Saltando.")
                    continue

                # Si CUALQUIERA de las unidades está NO_OPERATIVA, no generar alertas ambientales para ese aire.
                if aire_actual.evaporadora_operativa == OperativaStateEnum.NO_OPERATIVA or \
                   aire_actual.condensadora_operativa == OperativaStateEnum.NO_OPERATIVA:
                    print(f"[DEBUG] Aire ID: {lectura.aire_id} tiene al menos una unidad NO OPERATIVA, saltando verificación de umbrales ambientales.")
                    continue
                
                print(f"\n[DEBUG] Verificando (ambiental) Aire ID: {lectura.aire_id} (T:{lectura.temperatura}, H:{lectura.humedad})")
                alerta_ambiental_encontrada = False
                umbrales_specific_para_aire = specific_umbrales_dict.get(lectura.aire_id, [])
                umbrales_aplicables_para_aire = global_umbrales + umbrales_specific_para_aire

                if not umbrales_aplicables_para_aire:
                    print("[DEBUG]   No hay umbrales ambientales aplicables para este aire.")
                    continue

                for umbral in umbrales_aplicables_para_aire:
                    print(f"[DEBUG]     Comparando con Umbral ID: {umbral.id} (T:[{umbral.temp_min},{umbral.temp_max}], H:[{umbral.hum_min},{umbral.hum_max}])")
                    try:
                        temp_lectura = float(lectura.temperatura)
                        hum_lectura = float(lectura.humedad) if lectura.humedad is not None else None
                        temp_min = float(umbral.temp_min)
                        temp_max = float(umbral.temp_max)
                        hum_min = float(umbral.hum_min)
                        hum_max = float(umbral.hum_max)

                        fuera_limite_temp = (temp_lectura < temp_min or temp_lectura > temp_max)
                        fuera_limite_hum = (hum_lectura is not None and (hum_lectura < hum_min or hum_lectura > hum_max))
                        print(f"[DEBUG]       Temp fuera: {fuera_limite_temp} ({temp_lectura} vs [{temp_min}, {temp_max}])")
                        print(f"[DEBUG]       Hum fuera: {fuera_limite_hum} ({hum_lectura} vs [{hum_min}, {hum_max}])")

                        if fuera_limite_temp or fuera_limite_hum:
                            print(f"[DEBUG]       ¡VIOLACIÓN AMBIENTAL DETECTADA por umbral {umbral.id} para Aire ID {lectura.aire_id}!")
                            aires_con_alerta.add(lectura.aire_id)
                            alerta_ambiental_encontrada = True
                            break 
                        else:
                             print(f"[DEBUG]       Dentro de límites para umbral {umbral.id}.")
                    except Exception as e_compare:
                        print(f"[ERROR-DEBUG] Error al comparar lectura ambiental con umbral {umbral.id}: {e_compare}")
                
                if not alerta_ambiental_encontrada:
                    print(f"[DEBUG]   Lectura ambiental DENTRO de todos los umbrales aplicables para Aire ID {lectura.aire_id}.")

        total_alertas = len(aires_con_alerta)
        print(f"\n--- [DEBUG] Fin contar_alertas_activas_helper. Total aires con alguna alerta: {total_alertas} ---")
        return total_alertas

    except Exception as e:
        print(f"!!! ERROR GENERAL en contar_alertas_activas_helper: {e}", file=sys.stderr)
        traceback.print_exc()
        return -1 # Indicar error con -1

def obtener_ultimas_lecturas_con_info_aire_helper(limite=5):
    """
    Helper para obtener las últimas N lecturas con info del aire.
    Retorna una lista de diccionarios o None en caso de error.
    """
    try:
        # Alias para claridad
        LecturaAlias = aliased(Lectura)
        AireAlias = aliased(AireAcondicionado)

        # Consulta
        query = db.session.query(
            LecturaAlias.id,
            LecturaAlias.aire_id,
            AireAlias.nombre.label('nombre_aire'),
            AireAlias.ubicacion.label('ubicacion_aire'),
            LecturaAlias.temperatura,
            LecturaAlias.humedad,
            LecturaAlias.fecha
        ).join(
            AireAlias, LecturaAlias.aire_id == AireAlias.id
        ).order_by(
            desc(LecturaAlias.fecha)
        ).limit(limite)

        # Ejecutar y convertir a lista de diccionarios
        ultimas_lecturas_raw = query.all()

        # Convertir resultados (NamedTuple) a diccionarios serializables
        results = []
        for lectura in ultimas_lecturas_raw:
            results.append({
                'id': lectura.id,
                'aire_id': lectura.aire_id,
                'nombre_aire': lectura.nombre_aire,
                'ubicacion_aire': lectura.ubicacion_aire,
                'temperatura': lectura.temperatura,
                'humedad': lectura.humedad,
                # Formatear fecha a ISO string para JSON
                'fecha': lectura.fecha.isoformat() if lectura.fecha else None
            })

        return results

    except Exception as e:
        print(f"Error al obtener últimas lecturas con info aire: {e}", file=sys.stderr)
        traceback.print_exc()
        return None # Indicar error

@api.route('/lecturas/ubicacion/<path:ubicacion>', methods=['GET']) # Usar <path:> para capturar slashes en la ubicación
@jwt_required()
def obtener_lecturas_por_ubicacion_route(ubicacion):
    """
    Obtiene las últimas N lecturas para una ubicación específica.
    Requiere autenticación.
    Acepta query param 'limite' (default 50).
    """
    try:
        limite = request.args.get('limite', default=50, type=int)
        if limite <= 0 or limite > 500: # Limitar a 500 por rendimiento
            limite = 50
    except ValueError:
        limite = 50

    try:
        # Buscar los IDs de los aires en esa ubicación
        aires_en_ubicacion = db.session.query(AireAcondicionado.id)\
            .filter(AireAcondicionado.ubicacion == ubicacion)\
            .all()

        if not aires_en_ubicacion:
            return jsonify([]), 200 # No hay aires, devuelve lista vacía

        aire_ids = [a.id for a in aires_en_ubicacion]

        # Obtener las últimas 'limite' lecturas para esos aires
        lecturas = db.session.query(Lectura)\
            .filter(Lectura.aire_id.in_(aire_ids))\
            .order_by(Lectura.fecha.desc())\
            .limit(limite)\
            .all()

        # Ordenar por fecha ascendente para el gráfico
        lecturas.reverse()

        return jsonify([l.serialize() for l in lecturas]), 200

    except Exception as e:
        print(f"Error obteniendo lecturas para ubicación '{ubicacion}': {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error al obtener lecturas por ubicación."}), 500
# --- Rutas de API ---

@api.route('/contadores', methods=['GET'])
@jwt_required() 
def obtener_contadores_route():
    """
    Endpoint para obtener los contadores totales de aires, lecturas,
    mantenimientos y otros equipos. Requiere autenticación.
    """
    # No se necesita get_jwt_identity() aquí, solo autenticación
    counts = contar_entidades_helper()
    # Verificar si algún contador dio error (-1)
   # if any(v == -1 for v in counts.values()): # Comentado temporalmente si -1 es un valor válido para algún contador
    #      return jsonify({"msg": "Error al obtener algunos contadores.", "counts": counts}), 500
 
    return jsonify(counts), 200

@api.route('/alertas/activas/count', methods=['GET'])
@jwt_required() # <--- Añadido aquí
def obtener_contador_alertas_activas_route():
    """
    Endpoint para obtener el número de aires con alertas activas
    basado en su última lectura. Requiere autenticación.
    """
    # No se necesita get_jwt_identity() aquí, solo autenticación
    alert_count = contar_alertas_activas_helper()
    if alert_count == -1:
        return jsonify({"msg": "Error al calcular el número de alertas activas."}), 500
    return jsonify({"alertas_activas_count": alert_count}), 200

@api.route('/lecturas/ultimas', methods=['GET'])
@jwt_required() 
def obtener_ultimas_lecturas_route():
    """
    Endpoint para obtener las últimas N lecturas registradas,
    incluyendo información del aire asociado. Requiere autenticación.
    Acepta un query parameter 'limite' (default 5).
    """
    try:
        limite = request.args.get('limite', default=5, type=int)
        if limite <= 0 or limite > 100:
            limite = 5
    except ValueError:
        limite = 5

    # Llama al helper que ya tenías (asumiendo que está definido correctamente)
    ultimas_lecturas = obtener_ultimas_lecturas_con_info_aire_helper(limite)

    if ultimas_lecturas is None:
        return jsonify({"msg": "Error al obtener las últimas lecturas."}), 500

    return jsonify(ultimas_lecturas), 200

##Proveedores

@api.route('/proveedores', methods=['POST'])
@jwt_required()
def crear_proveedor_route():
    """Crea un nuevo proveedor."""
    current_user_id = get_jwt_identity()
    logged_in_user = db.session.get(TrackerUsuario, current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor']:
        return jsonify({"msg": "Acceso no autorizado"}), 403

    data = request.get_json()
    if not data or not data.get('nombre'):
        return jsonify({"msg": "El campo 'nombre' es requerido."}), 400

    try:
        nuevo_proveedor = Proveedor(
            nombre=data['nombre'],
            email_proveedor=data.get('email_proveedor')
        )
        db.session.add(nuevo_proveedor)
        db.session.commit()
        return jsonify(nuevo_proveedor.serialize()), 201
    except ValueError as ve: # Captura validaciones del modelo
        db.session.rollback()
        return jsonify({"msg": str(ve)}), 400
    except IntegrityError: # Captura violación de unicidad (nombre)
        db.session.rollback()
        return jsonify({"msg": f"Ya existe un proveedor con el nombre '{data['nombre']}'."}), 409
    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"Error DB creando proveedor: {e}", file=sys.stderr)
        return jsonify({"msg": "Error de base de datos al crear proveedor."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"Error inesperado creando proveedor: {e}", file=sys.stderr)
        return jsonify({"msg": "Error inesperado en el servidor."}), 500

@api.route('/proveedores', methods=['GET'])
@jwt_required()
def obtener_proveedores_route():
    """Obtiene la lista de todos los proveedores."""
    try:
        proveedores = db.session.query(Proveedor).order_by(Proveedor.nombre).all()
        return jsonify([p.serialize() for p in proveedores]), 200
    except Exception as e:
        print(f"Error obteniendo proveedores: {e}", file=sys.stderr)
        return jsonify({"msg": "Error al obtener proveedores."}), 500

@api.route('/proveedores/<int:proveedor_id>', methods=['GET'])
@jwt_required()
def obtener_proveedor_por_id_route(proveedor_id):
    """Obtiene un proveedor específico por ID."""
    try:
        proveedor = db.session.get(Proveedor, proveedor_id)
        if not proveedor:
            return jsonify({"msg": "Proveedor no encontrado."}), 404
        return jsonify(proveedor.serialize()), 200
    except Exception as e:
        print(f"Error obteniendo proveedor {proveedor_id}: {e}", file=sys.stderr)
        return jsonify({"msg": "Error al obtener proveedor."}), 500

@api.route('/proveedores/<int:proveedor_id>', methods=['PUT'])
@jwt_required()
def actualizar_proveedor_route(proveedor_id):
    """Actualiza un proveedor existente."""
    current_user_id = get_jwt_identity()
    logged_in_user = db.session.get(TrackerUsuario, current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor']:
        return jsonify({"msg": "Acceso no autorizado"}), 403

    proveedor = db.session.get(Proveedor, proveedor_id)
    if not proveedor:
        return jsonify({"msg": "Proveedor no encontrado."}), 404

    data = request.get_json()
    if not data:
        return jsonify({"msg": "No se recibieron datos."}), 400

    updated = False
    try:
        if 'nombre' in data and data['nombre'] != proveedor.nombre:
            proveedor.nombre = data['nombre'] # La validación se dispara aquí
            updated = True
        # Permitir actualizar a None o vacío si se envía explícitamente
        if 'email_proveedor' in data and data['email_proveedor'] != proveedor.email_proveedor:
             proveedor.email_proveedor = data['email_proveedor'] if data['email_proveedor'] else None
             updated = True

        if updated:
            db.session.commit()
            return jsonify(proveedor.serialize()), 200
        else:
            return jsonify(proveedor.serialize()), 200 # O 304 Not Modified

    except ValueError as ve:
        db.session.rollback()
        return jsonify({"msg": str(ve)}), 400
    except IntegrityError:
        db.session.rollback()
        return jsonify({"msg": f"Ya existe otro proveedor con el nombre '{data.get('nombre')}'."}), 409
    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"Error DB actualizando proveedor {proveedor_id}: {e}", file=sys.stderr)
        return jsonify({"msg": "Error de base de datos."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"Error inesperado actualizando proveedor {proveedor_id}: {e}", file=sys.stderr)
        return jsonify({"msg": "Error inesperado."}), 500

@api.route('/proveedores/<int:proveedor_id>', methods=['DELETE'])
@jwt_required()
def eliminar_proveedor_route(proveedor_id):
    """Elimina un proveedor y sus contactos asociados."""
    current_user_id = get_jwt_identity()
    logged_in_user = db.session.get(TrackerUsuario, current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin']: # Solo Admin puede borrar proveedores
        return jsonify({"msg": "Acceso no autorizado"}), 403

    proveedor = db.session.get(Proveedor, proveedor_id)
    if not proveedor:
        return jsonify({"msg": "Proveedor no encontrado."}), 404

    try:
        db.session.delete(proveedor) # Cascade debería eliminar contactos
        db.session.commit()
        return '', 204
    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"Error DB eliminando proveedor {proveedor_id}: {e}", file=sys.stderr)
        return jsonify({"msg": "Error de base de datos."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"Error inesperado eliminando proveedor {proveedor_id}: {e}", file=sys.stderr)
        return jsonify({"msg": "Error inesperado."}), 500

# --- Rutas para Contactos de Proveedor ---

@api.route('/proveedores/<int:proveedor_id>/contactos', methods=['POST'])
@jwt_required()
def crear_contacto_route(proveedor_id):
    """Crea un nuevo contacto para un proveedor específico."""
    current_user_id = get_jwt_identity()
    logged_in_user = db.session.get(TrackerUsuario, current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor']:
        return jsonify({"msg": "Acceso no autorizado"}), 403

    proveedor = db.session.get(Proveedor, proveedor_id)
    if not proveedor:
        return jsonify({"msg": "Proveedor no encontrado."}), 404

    data = request.get_json()
    if not data or not data.get('nombre_contacto'):
        return jsonify({"msg": "El campo 'nombre_contacto' es requerido."}), 400

    try:
        nuevo_contacto = ContactoProveedor(
            proveedor_id=proveedor_id,
            nombre_contacto=data['nombre_contacto'],
            telefono_contacto=data.get('telefono_contacto'),
            email_contacto=data.get('email_contacto')
        )
        db.session.add(nuevo_contacto)
        db.session.commit()
        return jsonify(nuevo_contacto.serialize()), 201
    except ValueError as ve:
        db.session.rollback()
        return jsonify({"msg": str(ve)}), 400
    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"Error DB creando contacto para proveedor {proveedor_id}: {e}", file=sys.stderr)
        return jsonify({"msg": "Error de base de datos."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"Error inesperado creando contacto: {e}", file=sys.stderr)
        return jsonify({"msg": "Error inesperado."}), 500

@api.route('/proveedores/<int:proveedor_id>/contactos', methods=['GET'])
@jwt_required()
def obtener_contactos_por_proveedor_route(proveedor_id):
    """Obtiene los contactos de un proveedor específico."""
    proveedor = db.session.get(Proveedor, proveedor_id)
    if not proveedor:
        return jsonify({"msg": "Proveedor no encontrado."}), 404

    try:
        # Accede a los contactos a través de la relación
        contactos = proveedor.contactos
        return jsonify([c.serialize() for c in contactos]), 200
    except Exception as e:
        print(f"Error obteniendo contactos para proveedor {proveedor_id}: {e}", file=sys.stderr)
        return jsonify({"msg": "Error al obtener contactos."}), 500

@api.route('/contactos_proveedor/<int:contacto_id>', methods=['GET'])
@jwt_required()
def obtener_contacto_por_id_route(contacto_id):
    """Obtiene un contacto específico por su ID."""
    try:
        contacto = db.session.get(ContactoProveedor, contacto_id)
        if not contacto:
            return jsonify({"msg": "Contacto no encontrado."}), 404
        return jsonify(contacto.serialize()), 200
    except Exception as e:
        print(f"Error obteniendo contacto {contacto_id}: {e}", file=sys.stderr)
        return jsonify({"msg": "Error al obtener contacto."}), 500

@api.route('/contactos_proveedor/<int:contacto_id>', methods=['PUT'])
@jwt_required()
def actualizar_contacto_route(contacto_id):
    """Actualiza un contacto existente."""
    current_user_id = get_jwt_identity()
    logged_in_user = db.session.get(TrackerUsuario, current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor']:
        return jsonify({"msg": "Acceso no autorizado"}), 403

    contacto = db.session.get(ContactoProveedor, contacto_id)
    if not contacto:
        return jsonify({"msg": "Contacto no encontrado."}), 404

    data = request.get_json()
    if not data:
        return jsonify({"msg": "No se recibieron datos."}), 400

    updated = False
    try:
        if 'nombre_contacto' in data and data['nombre_contacto'] != contacto.nombre_contacto:
            contacto.nombre_contacto = data['nombre_contacto'] # Validación se dispara
            updated = True
        if 'telefono_contacto' in data and data['telefono_contacto'] != contacto.telefono_contacto:
            contacto.telefono_contacto = data['telefono_contacto'] if data['telefono_contacto'] else None
            updated = True
        if 'email_contacto' in data and data['email_contacto'] != contacto.email_contacto:
            contacto.email_contacto = data['email_contacto'] if data['email_contacto'] else None
            updated = True

        if updated:
            db.session.commit()
            return jsonify(contacto.serialize()), 200
        else:
            return jsonify(contacto.serialize()), 200 # O 304

    except ValueError as ve:
        db.session.rollback()
        return jsonify({"msg": str(ve)}), 400
    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"Error DB actualizando contacto {contacto_id}: {e}", file=sys.stderr)
        return jsonify({"msg": "Error de base de datos."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"Error inesperado actualizando contacto {contacto_id}: {e}", file=sys.stderr)
        return jsonify({"msg": "Error inesperado."}), 500

@api.route('/contactos_proveedor/<int:contacto_id>', methods=['DELETE'])
@jwt_required()
def eliminar_contacto_route(contacto_id):
    """Elimina un contacto específico."""
    current_user_id = get_jwt_identity()
    logged_in_user = db.session.get(TrackerUsuario, current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor']:
        return jsonify({"msg": "Acceso no autorizado"}), 403

    contacto = db.session.get(ContactoProveedor, contacto_id)
    if not contacto:
        return jsonify({"msg": "Contacto no encontrado."}), 404

    try:
        db.session.delete(contacto)
        db.session.commit()
        return '', 204
    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"Error DB eliminando contacto {contacto_id}: {e}", file=sys.stderr)
        return jsonify({"msg": "Error de base de datos."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"Error inesperado eliminando contacto {contacto_id}: {e}", file=sys.stderr)
        return jsonify({"msg": "Error inesperado."}), 500

@api.route('/proveedores/<int:proveedor_id>/actividades', methods=['POST'])
@jwt_required()
def crear_actividad_proveedor_route(proveedor_id):
    """Crea una nueva actividad para un proveedor específico."""
    current_user_id = get_jwt_identity()
    logged_in_user = db.session.get(TrackerUsuario, current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor']:
        return jsonify({"msg": "Acceso no autorizado para crear actividades"}), 403

    proveedor = db.session.get(Proveedor, proveedor_id)
    if not proveedor:
        return jsonify({"msg": "Proveedor no encontrado."}), 404

    data = request.get_json()
    if not data:
        return jsonify({"msg": "No se recibieron datos JSON."}), 400

    required_fields = ['descripcion', 'fecha_ocurrencia']
    if not all(field in data for field in required_fields):
        missing = [field for field in required_fields if field not in data]
        return jsonify({"msg": f"Faltan campos requeridos: {', '.join(missing)}"}), 400

    try:
        # Convertir fechas de string a datetime
        fecha_ocurrencia_dt = datetime.fromisoformat(data['fecha_ocurrencia'])
        fecha_reporte_dt = datetime.fromisoformat(data['fecha_reporte']) if 'fecha_reporte' in data and data['fecha_reporte'] else datetime.now(timezone.utc)


        # Validar y obtener estatus (default es Pendiente)
        estatus_val = data.get('estatus', EstatusActividad.PENDIENTE.value)
        try:
            estatus_enum = EstatusActividad(estatus_val)
        except ValueError:
            return jsonify({"msg": f"Valor de estatus inválido: {estatus_val}. Usar Pendiente, En Progreso, Completado o Cancelado."}), 400

        nueva_actividad = ActividadProveedor(
            proveedor_id=proveedor_id,
            descripcion=data['descripcion'],
            fecha_ocurrencia=fecha_ocurrencia_dt,
            fecha_reporte=fecha_reporte_dt,
            numero_reporte=data.get('numero_reporte'),
            estatus=estatus_enum
        )
        db.session.add(nueva_actividad)
        db.session.commit()
        return jsonify(nueva_actividad.serialize()), 201

    except (ValueError, TypeError) as ve: # Captura errores de conversión de fecha/enum o validación del modelo
        db.session.rollback()
        return jsonify({"msg": str(ve)}), 400
    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"Error DB creando actividad para proveedor {proveedor_id}: {e}", file=sys.stderr)
        return jsonify({"msg": "Error de base de datos al crear la actividad."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"Error inesperado creando actividad: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor."}), 500

@api.route('/proveedores/<int:proveedor_id>/actividades', methods=['GET'])
@jwt_required()
def obtener_actividades_por_proveedor_route(proveedor_id):
    """Obtiene las actividades de un proveedor específico."""
    proveedor = db.session.get(Proveedor, proveedor_id)
    if not proveedor:
        return jsonify({"msg": "Proveedor no encontrado."}), 404

    try:
        # Ordenar por fecha de ocurrencia descendente por defecto
        actividades = db.session.query(ActividadProveedor)\
            .filter_by(proveedor_id=proveedor_id)\
            .order_by(ActividadProveedor.fecha_ocurrencia.desc())\
            .all()
        return jsonify([a.serialize() for a in actividades]), 200
    except Exception as e:
        print(f"Error obteniendo actividades para proveedor {proveedor_id}: {e}", file=sys.stderr)
        return jsonify({"msg": "Error al obtener actividades."}), 500

@api.route('/actividades_proveedor', methods=['GET'])
@jwt_required()
def obtener_todas_actividades_route():
    """Obtiene todas las actividades, opcionalmente filtradas por estatus."""
    try:
        query = db.session.query(ActividadProveedor)
        estatus_filter = request.args.get('estatus')
        if estatus_filter:
            try:
                estatus_enum = EstatusActividad(estatus_filter)
                query = query.filter(ActividadProveedor.estatus == estatus_enum)
            except ValueError:
                return jsonify({"msg": f"Valor de estatus inválido para filtrar: {estatus_filter}"}), 400

        # Ordenar por fecha de reporte descendente
        actividades = query.order_by(ActividadProveedor.fecha_reporte.desc()).all()
        return jsonify([a.serialize() for a in actividades]), 200
    except Exception as e:
        print(f"Error obteniendo todas las actividades: {e}", file=sys.stderr)
        return jsonify({"msg": "Error al obtener todas las actividades."}), 500


@api.route('/actividades_proveedor/<int:actividad_id>', methods=['GET'])
@jwt_required()
def obtener_actividad_por_id_route(actividad_id):
    """Obtiene una actividad específica por su ID."""
    try:
        actividad = db.session.get(ActividadProveedor, actividad_id)
        if not actividad:
            return jsonify({"msg": "Actividad no encontrada."}), 404
        return jsonify(actividad.serialize()), 200
    except Exception as e:
        print(f"Error obteniendo actividad {actividad_id}: {e}", file=sys.stderr)
        return jsonify({"msg": "Error al obtener la actividad."}), 500

@api.route('/actividades_proveedor/<int:actividad_id>', methods=['PUT'])
@jwt_required()
def actualizar_actividad_route(actividad_id):
    """Actualiza una actividad existente (descripción, fechas, reporte, estatus)."""
    current_user_id = get_jwt_identity()
    logged_in_user = db.session.get(TrackerUsuario, current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor']:
        return jsonify({"msg": "Acceso no autorizado para actualizar actividades"}), 403

    actividad = db.session.get(ActividadProveedor, actividad_id)
    if not actividad:
        return jsonify({"msg": "Actividad no encontrada."}), 404

    data = request.get_json()
    if not data:
        return jsonify({"msg": "No se recibieron datos JSON."}), 400

    updated = False
    try:
        if 'descripcion' in data and data['descripcion'] != actividad.descripcion:
            actividad.descripcion = data['descripcion'] # Validación se dispara
            updated = True
        if 'fecha_ocurrencia' in data and data['fecha_ocurrencia']:
            try:
                new_fecha_ocurrencia = datetime.fromisoformat(data['fecha_ocurrencia'])
                if new_fecha_ocurrencia != actividad.fecha_ocurrencia:
                    actividad.fecha_ocurrencia = new_fecha_ocurrencia
                    updated = True
            except (ValueError, TypeError):
                 return jsonify({"msg": "Formato de fecha_ocurrencia inválido."}), 400
        if 'fecha_reporte' in data and data['fecha_reporte']:
             try:
                 new_fecha_reporte = datetime.fromisoformat(data['fecha_reporte'])
                 if new_fecha_reporte != actividad.fecha_reporte:
                     actividad.fecha_reporte = new_fecha_reporte
                     updated = True
             except (ValueError, TypeError):
                  return jsonify({"msg": "Formato de fecha_reporte inválido."}), 400
        # Permitir actualizar numero_reporte a None o vacío
        if 'numero_reporte' in data and data['numero_reporte'] != actividad.numero_reporte:
             actividad.numero_reporte = data['numero_reporte'] if data['numero_reporte'] else None
             updated = True
        if 'estatus' in data:
            try:
                new_estatus_enum = EstatusActividad(data['estatus'])
                if new_estatus_enum != actividad.estatus:
                    actividad.estatus = new_estatus_enum
                    updated = True
            except ValueError:
                 return jsonify({"msg": f"Valor de estatus inválido: {data['estatus']}"}), 400

        if updated:
            actividad.ultima_modificacion = datetime.utcnow() # Actualizar timestamp
            db.session.commit()
            return jsonify(actividad.serialize()), 200
        else:
            return jsonify(actividad.serialize()), 200 # O 304 Not Modified

    except ValueError as ve: # Captura validaciones del modelo
        db.session.rollback()
        return jsonify({"msg": str(ve)}), 400
    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"Error DB actualizando actividad {actividad_id}: {e}", file=sys.stderr)
        return jsonify({"msg": "Error de base de datos al actualizar la actividad."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"Error inesperado actualizando actividad {actividad_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado en el servidor."}), 500

@api.route('/actividades_proveedor/<int:actividad_id>', methods=['DELETE'])
@jwt_required()
def eliminar_actividad_route(actividad_id):
    """Elimina una actividad específica."""
    current_user_id = get_jwt_identity()
    logged_in_user = db.session.get(TrackerUsuario, current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor']: # O solo admin si prefieres
        return jsonify({"msg": "Acceso no autorizado para eliminar actividades"}), 403

    actividad = db.session.get(ActividadProveedor, actividad_id)
    if not actividad:
        return jsonify({"msg": "Actividad no encontrada."}), 404

    try:
        db.session.delete(actividad)
        db.session.commit()
        return '', 204
    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"Error DB eliminando actividad {actividad_id}: {e}", file=sys.stderr)
        return jsonify({"msg": "Error de base de datos al eliminar la actividad."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"Error inesperado eliminando actividad {actividad_id}: {e}", file=sys.stderr)
        return jsonify({"msg": "Error inesperado en el servidor."}), 500

##Documentos

def allowed_file(filename):
    """Verifica si la extensión del archivo es permitida."""
    # --- Mantenemos la validación de extensión por seguridad ---
    allowed_extensions = current_app.config.get('ALLOWED_EXTENSIONS', {'txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif', 'doc', 'docx', 'xls', 'xlsx'})
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in allowed_extensions
    # ---------------------------------------------------------

# --- Rutas para Documentos Externos  ---

@api.route('/documentos', methods=['POST'])
@jwt_required()
def subir_documento_route():
    """Sube un nuevo documento externo (guardado en DB)."""
    current_user_id_str = get_jwt_identity()
    try:
        current_user_id = int(current_user_id_str)
    except ValueError:
        return jsonify({"msg": "Identidad de usuario inválida"}), 400

    logged_in_user = db.session.get(TrackerUsuario, current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor']:
        return jsonify({"msg": "Acceso no autorizado para subir documentos"}), 403

    if 'file' not in request.files:
        return jsonify({"msg": "No se encontró el archivo en la solicitud ('file')"}), 400

    file = request.files['file']
    nombre_descriptivo = request.form.get('nombre')
    descripcion = request.form.get('descripcion')

    if not nombre_descriptivo:
        return jsonify({"msg": "El campo 'nombre' es requerido en el formulario."}), 400

    if file.filename == '':
        return jsonify({"msg": "No se seleccionó ningún archivo."}), 400

    if file and allowed_file(file.filename):
        original_filename = secure_filename(file.filename)
        # --- Leer contenido del archivo ---
        try:
            file_content = file.read()
            if not file_content:
                 return jsonify({"msg": "El archivo parece estar vacío."}), 400
        except Exception as read_error:
            print(f"Error leyendo el archivo: {read_error}", file=sys.stderr)
            return jsonify({"msg": "Error al leer el contenido del archivo."}), 500
        # ----------------------------------

        try:
            nuevo_documento = DocumentoExterno(
                nombre=nombre_descriptivo,
                descripcion=descripcion,
                nombre_archivo_original=original_filename,
                # --- Guardar contenido en DB ---
                datos_archivo=file_content,
                # -------------------------------
                tipo_mime=file.mimetype,
                usuario_carga_id=current_user_id
            )
            db.session.add(nuevo_documento)
            db.session.commit()

            return jsonify(nuevo_documento.serialize()), 201

        except SQLAlchemyError as db_error:
            db.session.rollback()
            print(f"Error DB al guardar documento: {db_error}", file=sys.stderr)
            return jsonify({"msg": "Error de base de datos al guardar el documento."}), 500
        except Exception as e:
            db.session.rollback()
            print(f"Error subiendo archivo a DB: {e}", file=sys.stderr)
            traceback.print_exc()
            return jsonify({"msg": "Error al guardar el archivo en la base de datos."}), 500
    else:
        return jsonify({"msg": "Tipo de archivo no permitido."}), 400


@api.route('/documentos', methods=['GET'])
@jwt_required()
def listar_documentos_route():
    """Obtiene la lista de todos los documentos externos."""
    try:
        documentos = db.session.query(DocumentoExterno)\
            .order_by(DocumentoExterno.fecha_carga.desc())\
            .all()

        results = []
        for doc in documentos:
            doc_data = doc.serialize()
            # La URL de descarga sigue apuntando a la ruta de descarga
            doc_data['url_descarga'] = url_for('api.descargar_documento_route',
                                               documento_id=doc.id,
                                               _external=True)
            results.append(doc_data)

        return jsonify(results), 200
    except Exception as e:
        print(f"Error listando documentos: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error al obtener la lista de documentos."}), 500


@api.route('/documentos/<int:documento_id>/download', methods=['GET'])
@jwt_required()
def descargar_documento_route(documento_id):
    """Descarga el archivo de un documento específico (desde DB)."""
    try:
        documento = db.session.get(DocumentoExterno, documento_id)
        if not documento:
            return jsonify({"msg": "Documento no encontrado."}), 404

        # --- Verificar si hay datos ---
        if not documento.datos_archivo:
             print(f"Advertencia: Documento ID {documento_id} no tiene datos_archivo en DB.", file=sys.stderr)
             return jsonify({"msg": "El archivo no tiene contenido almacenado."}), 404
        # -----------------------------

        # --- Usar send_file con BytesIO ---
        return send_file(
            io.BytesIO(documento.datos_archivo),
            mimetype=documento.tipo_mime or 'application/octet-stream', # Default MIME type
            as_attachment=True, # Forzar descarga
            download_name=documento.nombre_archivo_original # Nombre que verá el usuario
        )
        # ----------------------------------

    except Exception as e:
        print(f"Error descargando documento {documento_id} desde DB: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error al descargar el archivo."}), 500


@api.route('/documentos/<int:documento_id>', methods=['DELETE'])
@jwt_required()
def eliminar_documento_route(documento_id):
    """Elimina un documento (solo registro en DB)."""
    current_user_id_str = get_jwt_identity()
    try:
        current_user_id = int(current_user_id_str)
    except ValueError:
        return jsonify({"msg": "Identidad de usuario inválida"}), 400

    logged_in_user = db.session.get(TrackerUsuario, current_user_id)
    if not logged_in_user or logged_in_user.rol not in ['admin', 'supervisor']:
        return jsonify({"msg": "Acceso no autorizado para eliminar documentos"}), 403

    documento = db.session.get(DocumentoExterno, documento_id)
    if not documento:
        return jsonify({"msg": "Documento no encontrado."}), 404

    try:
        db.session.delete(documento)
        db.session.commit()

        return '', 204 # No Content

    except SQLAlchemyError as db_error:
        db.session.rollback()
        print(f"Error DB al eliminar documento {documento_id}: {db_error}", file=sys.stderr)
        return jsonify({"msg": "Error de base de datos al eliminar el documento."}), 500
    except Exception as e:
        db.session.rollback()
        print(f"Error inesperado eliminando documento {documento_id}: {e}", file=sys.stderr)
        traceback.print_exc()
        return jsonify({"msg": "Error inesperado al eliminar el documento."}), 500

@api.route('/alertas_activas_detalladas', methods=['GET'])
@jwt_required()
def get_alertas_activas_detalladas_route():
    """
    Endpoint para obtener un listado detallado de todas las alertas activas,
    incluyendo aires no operativos y alertas de temperatura/humedad.
    Requiere autenticación.
    """
    try:
        detalles = obtener_detalles_alertas_activas_helper()
        return jsonify(detalles), 200
    except Exception as e:
        print(f"Error en get_alertas_activas_detalladas_route: {e}", file=sys.stderr)
        return jsonify({"msg": "Error al obtener detalles de alertas activas."}), 500
